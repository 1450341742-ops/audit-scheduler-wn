import csv
import io
import json
import os
import re
import hashlib
from contextlib import contextmanager
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

import pandas as pd
import streamlit as st
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import text
from sqlalchemy.exc import IntegrityError

from app.db import Base, SessionLocal, engine, ensure_schema, IS_SQLITE, IS_SQLITE
from app.models import Auditor, Task, Schedule, CityDistance, City
from app.scheduler import (
    build_candidates,
    propose_team,
    compute_from_city,
    get_distance_km,
    team_objective,
)
from app.seed_distances import SEED_CITY_DISTANCES, CITY_COORDS

APP_NAME = "万宁睿和稽查排班"
st.set_page_config(page_title=APP_NAME, layout="wide")

# -------------------- 上传控件中文化 --------------------
st.markdown(
    """
    <style>
    [data-testid="stFileUploaderDropzoneInstructions"]{
        display:none !important;
    }

    [data-testid="stFileUploaderDropzone"]{
        position: relative !important;
    }
    [data-testid="stFileUploaderDropzone"]::before{
        content:"将文件拖拽到此处，或点击右侧“浏览文件”上传（支持 .xlsx/.csv，单个文件 ≤200MB）";
        display:block;
        color:#333;
        padding:10px 6px 8px 6px;
        line-height:1.5;
        font-size:14px;
        white-space:normal;
    }

    [data-testid="stFileUploaderDropzone"] input[type="file"]{
        width: 100% !important;
    }

    [data-testid="stFileUploaderDropzone"] input[type="file"]::file-selector-button{
        min-width: 132px !important;
        height: 42px !important;
        padding: 0 18px !important;
        border-radius: 10px !important;
        color: transparent !important;
        -webkit-text-fill-color: transparent !important;
        text-shadow: none !important;
    }

    [data-testid="stFileUploaderDropzone"]::after{
        content:"浏览文件";
        position:absolute;
        right: 14px;
        top: 50%;
        transform: translateY(-50%);
        min-width: 132px;
        height: 42px;
        padding: 0 18px;
        border-radius: 10px;
        border: 1px solid rgba(0,0,0,0.15);
        background: rgba(255,255,255,0.96);
        display:flex;
        align-items:center;
        justify-content:center;
        font-size:14px;
        font-weight:600;
        color:#111;
        white-space:nowrap;
        pointer-events:none;
        z-index: 9999;
        box-sizing: border-box;
    }

    @media (max-width: 520px){
        [data-testid="stFileUploaderDropzone"] input[type="file"]::file-selector-button{
            min-width:124px !important;
            height:40px !important;
            padding:0 14px !important;
        }
        [data-testid="stFileUploaderDropzone"]::after{
            min-width:124px;
            height:40px;
            padding:0 14px;
            right: 12px;
            font-size:14px;
        }
        [data-testid="stFileUploaderDropzone"]::before{
            font-size:13px;
        }
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# -------------------- 初始化 --------------------
@st.cache_resource(show_spinner=False)
def initialize_app_once():
    Base.metadata.create_all(bind=engine)
    ensure_schema()
    ensure_support_tables()
    with db_session() as db:
        seed_city_distances_if_needed(db)
        seed_cities_if_needed(db)
    bootstrap_auth_users_if_needed()
    return True


@contextmanager
def db_session():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def safe_parse_date(value) -> Optional[date]:
    if value is None:
        return None

    try:
        if pd.isna(value):
            return None
    except Exception:
        pass

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    try:
        if isinstance(value, pd.Timestamp):
            return value.to_pydatetime().date()
    except Exception:
        pass

    try:
        if isinstance(value, (int, float)) and not (isinstance(value, float) and pd.isna(value)):
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=float(value))).date()

        s_num = str(value).strip()
        if re.fullmatch(r"\d+(\.\d+)?", s_num):
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=float(s_num))).date()
    except Exception:
        pass

    s = str(value).strip()
    if not s:
        return None

    if " " in s:
        s = s.split(" ")[0].strip()
    s = s.replace("/", "-").replace(".", "-")

    for fmt in ("%Y-%m-%d", "%Y-%m", "%Y%m%d"):
        try:
            d = datetime.strptime(s, fmt).date()
            if fmt == "%Y-%m":
                return d.replace(day=1)
            return d
        except Exception:
            pass

    return None


def d2s(v: Optional[date]) -> str:
    return v.strftime("%Y-%m-%d") if v else ""


def show_table(rows: list[dict], height: int = 380):
    if not rows:
        st.info("暂无数据")
        return
    st.dataframe(rows, use_container_width=True, height=height)


def safe_commit(db: Session, context: str = "") -> bool:
    try:
        db.commit()
        return True
    except IntegrityError as e:
        db.rollback()
        st.error(f"数据库写入失败：{context}。常见原因：重复数据 / 唯一约束冲突。")
        st.exception(e)
        return False
    except Exception as e:
        db.rollback()
        st.error(f"数据库写入失败：{context}")
        st.exception(e)
        return False


def clear_runtime_caches_after_data_change():
    st.session_state["_data_version"] = int(st.session_state.get("_data_version", 0) or 0) + 1
    try:
        st.cache_data.clear()
    except Exception:
        pass
    for k in [
        "recommend_result",
        "batch_report",
        "smart_task_select",
        "member_ids_text",
        "edit_auditor_select",
        "edit_task_select",
    ]:
        if k in st.session_state:
            st.session_state.pop(k, None)


def get_data_version() -> int:
    return int(st.session_state.get("_data_version", 0) or 0)


def _safe_int(x, default=None):
    try:
        if x is None:
            return default
        if isinstance(x, float) and pd.isna(x):
            return default
        s = str(x).strip()
        if s == "":
            return default
        return int(float(s))
    except Exception:
        return default


def normalize_text(v) -> str:
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass
    return str(v).strip()


def seed_city_distances_if_needed(db: Session):
    seen = set()
    for a, b, km in SEED_CITY_DISTANCES:
        a = str(a).strip()
        b = str(b).strip()
        if not a or not b or a == b:
            continue
        key = (a, b)
        if key in seen:
            continue
        seen.add(key)
        exists = db.query(CityDistance).filter(CityDistance.from_city == a, CityDistance.to_city == b).first()
        if exists:
            continue
        db.add(CityDistance(from_city=a, to_city=b, km=float(km)))
        try:
            db.flush()
        except IntegrityError:
            db.rollback()
            continue
    safe_commit(db, "初始化城市距离")


SEED_CITIES = [(name, latlon[0], latlon[1]) for name, latlon in CITY_COORDS.items()]


def seed_cities_if_needed(db: Session):
    if db.query(City).count() > 0:
        return
    for name, lat, lon in SEED_CITIES:
        nm = str(name).strip()
        if not nm:
            continue
        db.add(City(name=nm, lat=float(lat), lon=float(lon)))
        try:
            db.flush()
        except IntegrityError:
            db.rollback()
            continue
    safe_commit(db, "初始化城市坐标")



# -------------------- 权限 --------------------
ALL_PAGES = [
    "智能排班",
    "批量排班",
    "稽查员管理",
    "任务管理",
    "指标统计",
    "兼职库",
    "城市距离",
    "城市坐标",
    "模板导入",
    "日历视图",
    "账号管理",
    "数据清理",
]
DEFAULT_NORMAL_PAGES = ["任务管理", "稽查员管理", "日历视图", "指标统计"]


def hash_password(password: str) -> str:
    return hashlib.sha256(str(password).encode("utf-8")).hexdigest()


def ensure_auth_table():
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS auth_users (
                    username TEXT PRIMARY KEY,
                    password_hash TEXT NOT NULL,
                    is_admin INTEGER NOT NULL DEFAULT 0,
                    is_super_admin INTEGER NOT NULL DEFAULT 0,
                    allowed_pages_json TEXT,
                    created_at TEXT
                )
                """
            )
        )

    if IS_SQLITE:
        with engine.begin() as conn:
            cols = conn.execute(text("PRAGMA table_info(auth_users)")).mappings().all()
            existing = {str(c.get("name")) for c in cols}

            if "is_super_admin" not in existing:
                conn.execute(text("ALTER TABLE auth_users ADD COLUMN is_super_admin INTEGER NOT NULL DEFAULT 0"))

            if "allowed_pages_json" not in existing:
                conn.execute(text("ALTER TABLE auth_users ADD COLUMN allowed_pages_json TEXT"))
    else:
        with engine.begin() as conn:
            conn.execute(
                text("ALTER TABLE auth_users ADD COLUMN IF NOT EXISTS is_super_admin INTEGER NOT NULL DEFAULT 0")
            )
            conn.execute(
                text("ALTER TABLE auth_users ADD COLUMN IF NOT EXISTS allowed_pages_json TEXT")
            )

def _bootstrap_seed_users() -> dict[str, str]:
    users = {}
    try:
        secret_users = st.secrets.get("auth_users", None)
        if secret_users:
            users = {str(k): str(v) for k, v in dict(secret_users).items()}
    except Exception:
        pass
    if not users:
        env_json = os.environ.get("AUTH_USERS_JSON", "").strip()
        if env_json:
            try:
                data = json.loads(env_json)
                if isinstance(data, dict):
                    users = {str(k): str(v) for k, v in data.items()}
            except Exception:
                pass
    if not users:
        users = {"admin": "admin123"}
    return users


def bootstrap_auth_users_if_needed():
    ensure_auth_table()
    with engine.begin() as conn:
        count = conn.execute(text("SELECT COUNT(*) FROM auth_users")).scalar() or 0
        if int(count) > 0:
            row = conn.execute(
                text("SELECT username, is_super_admin FROM auth_users WHERE username='admin'")
            ).mappings().first()
            if row and int(row.get("is_super_admin", 0)) != 1:
                conn.execute(text("UPDATE auth_users SET is_admin=1, is_super_admin=1 WHERE username='admin'"))
            return

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for username, password in _bootstrap_seed_users().items():
            clean_user = str(username).strip()
            if not clean_user:
                continue

            is_admin = 1 if clean_user == "admin" else 0
            is_super = 1 if clean_user == "admin" else 0
            allowed = None if is_admin else json.dumps(DEFAULT_NORMAL_PAGES, ensure_ascii=False)

            conn.execute(
                text(
                    """
                    INSERT INTO auth_users (username, password_hash, is_admin, is_super_admin, allowed_pages_json, created_at)
                    VALUES (:username, :password_hash, :is_admin, :is_super_admin, :allowed_pages_json, :created_at)
                    """
                ),
                {
                    "username": clean_user,
                    "password_hash": hash_password(str(password)),
                    "is_admin": is_admin,
                    "is_super_admin": is_super,
                    "allowed_pages_json": allowed,
                    "created_at": now,
                },
            )


def list_auth_users() -> list[dict]:
    ensure_auth_table()
    with engine.begin() as conn:
        rows = conn.execute(
            text(
                """
                SELECT username, is_admin, is_super_admin, allowed_pages_json, created_at
                FROM auth_users
                ORDER BY is_super_admin DESC, is_admin DESC, username ASC
                """
            )
        ).mappings().all()
    return [dict(r) for r in rows]


def get_auth_user(username: str) -> Optional[dict]:
    ensure_auth_table()
    clean_user = str(username or "").strip()
    if not clean_user:
        return None
    with engine.begin() as conn:
        row = conn.execute(
            text(
                """
                SELECT username, password_hash, is_admin, is_super_admin, allowed_pages_json, created_at
                FROM auth_users
                WHERE username = :username
                """
            ),
            {"username": clean_user},
        ).mappings().first()
    return dict(row) if row else None


def _normalize_pages(pages: list[str]) -> list[str]:
    seen = set()
    out = []
    for p in pages or []:
        p = str(p).strip()
        if not p:
            continue
        if p not in ALL_PAGES:
            continue
        if p in seen:
            continue
        seen.add(p)
        out.append(p)
    return out


def get_user_allowed_pages(username: str) -> list[str]:
    u = get_auth_user(username)
    if not u:
        return DEFAULT_NORMAL_PAGES[:]
    if int(u.get("is_admin", 0)) == 1:
        return ALL_PAGES[:]
    raw = u.get("allowed_pages_json") or ""
    try:
        arr = json.loads(raw) if raw else []
        if isinstance(arr, list):
            pages = _normalize_pages(arr)
            return pages if pages else DEFAULT_NORMAL_PAGES[:]
    except Exception:
        pass
    return DEFAULT_NORMAL_PAGES[:]


def set_user_allowed_pages(username: str, pages: list[str]) -> tuple[bool, str]:
    ensure_auth_table()
    clean_user = str(username or "").strip()
    if not clean_user:
        return False, "账号不能为空"
    u = get_auth_user(clean_user)
    if not u:
        return False, "账号不存在"
    if int(u.get("is_admin", 0)) == 1:
        return False, "管理员账号默认全功能，无需设置可见板块"
    pages = _normalize_pages(pages)
    if not pages:
        return False, "至少勾选1个可见板块"
    with engine.begin() as conn:
        conn.execute(
            text("UPDATE auth_users SET allowed_pages_json = :v WHERE username = :username"),
            {"v": json.dumps(pages, ensure_ascii=False), "username": clean_user},
        )
    return True, "已保存可见板块"


def create_auth_user(username: str, password: str, is_admin: bool = False, is_super_admin: bool = False) -> tuple[bool, str]:
    ensure_auth_table()
    clean_user = str(username or "").strip()
    if not clean_user:
        return False, "账号不能为空"
    if len(clean_user) < 3:
        return False, "账号至少3位"
    if not re.fullmatch(r"[A-Za-z0-9_.-]+", clean_user):
        return False, "账号仅支持字母、数字、下划线、点、短横线"
    if len(str(password or "")) < 6:
        return False, "密码至少6位"
    if get_auth_user(clean_user):
        return False, "该账号已存在"

    if is_super_admin:
        is_admin = True

    allowed = None if is_admin else json.dumps(DEFAULT_NORMAL_PAGES, ensure_ascii=False)

    with engine.begin() as conn:
        conn.execute(
            text(
                """
                INSERT INTO auth_users (username, password_hash, is_admin, is_super_admin, allowed_pages_json, created_at)
                VALUES (:username, :password_hash, :is_admin, :is_super_admin, :allowed_pages_json, :created_at)
                """
            ),
            {
                "username": clean_user,
                "password_hash": hash_password(password),
                "is_admin": 1 if is_admin else 0,
                "is_super_admin": 1 if is_super_admin else 0,
                "allowed_pages_json": allowed,
                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            },
        )
    return True, "新增账号成功"


def update_auth_password(username: str, new_password: str) -> tuple[bool, str]:
    ensure_auth_table()
    clean_user = str(username or "").strip()
    if not clean_user:
        return False, "账号不能为空"
    if len(str(new_password or "")) < 6:
        return False, "新密码至少6位"
    if not get_auth_user(clean_user):
        return False, "账号不存在"
    with engine.begin() as conn:
        conn.execute(
            text("UPDATE auth_users SET password_hash = :password_hash WHERE username = :username"),
            {"username": clean_user, "password_hash": hash_password(new_password)},
        )
    return True, "密码修改成功"


def delete_auth_user(username: str, current_user: str) -> tuple[bool, str]:
    ensure_auth_table()
    clean_user = str(username or "").strip()
    if clean_user == "admin":
        return False, "默认管理员 admin 不允许删除"
    if clean_user == str(current_user or "").strip():
        return False, "不能删除当前登录账号"
    if not get_auth_user(clean_user):
        return False, "账号不存在"
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM auth_users WHERE username = :username"), {"username": clean_user})
    return True, "账号已删除"


def check_login(username: str, password: str) -> bool:
    user = get_auth_user(username)
    if not user:
        return False
    return str(user.get("password_hash")) == hash_password(str(password))




def render_login():
    st.title(APP_NAME)
    st.subheader("账号密码登录")
    st.caption("首次使用默认主管理员：admin / admin123")
    with st.form("login_form", clear_on_submit=False):
        username = st.text_input("账号")
        password = st.text_input("密码", type="password")
        submitted = st.form_submit_button("登录", type="primary")
    if submitted:
        if check_login(username, password):
            user = get_auth_user(username)
            st.session_state["logged_in"] = True
            st.session_state["login_user"] = str(username).strip()
            st.session_state["is_admin"] = bool(int(user.get("is_admin", 0))) if user else False
            st.session_state["is_super_admin"] = bool(int(user.get("is_super_admin", 0))) if user else False
            st.session_state["allowed_pages"] = get_user_allowed_pages(str(username).strip())
            st.rerun()
        else:
            st.error("账号或密码错误")
    st.stop()


if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False
if "is_admin" not in st.session_state:
    st.session_state["is_admin"] = False
if "is_super_admin" not in st.session_state:
    st.session_state["is_super_admin"] = False
if "allowed_pages" not in st.session_state:
    st.session_state["allowed_pages"] = DEFAULT_NORMAL_PAGES[:]

if not st.session_state["logged_in"]:
    render_login()

STATUS_MAP = {"在岗": "active", "请假": "leave", "冻结": "frozen"}
STATUS_MAP_REV = {v: k for k, v in STATUS_MAP.items()}
BOOL_TRUE = {"是", "Y", "y", "yes", "YES", "True", "true", "1", "是/yes"}


def ics_escape(s: str) -> str:
    return (s or "").replace("\\", "\\\\").replace(";", "\\;").replace(",", "\\,").replace("\n", "\\n")


def build_ics_events(db: Session, auditor_id: int | None = None):
    q = db.query(Schedule).options(joinedload(Schedule.task), joinedload(Schedule.auditor)).order_by(Schedule.id.desc())
    if auditor_id:
        q = q.filter(Schedule.auditor_id == auditor_id)
    sch = q.all()
    events = []
    now = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    for s in sch:
        a = s.auditor
        t = s.task
        if not a or not t:
            continue
        start = datetime.combine(t.start_date, datetime.min.time()).replace(hour=9)
        actual_end = t.end_date or (t.start_date + timedelta(days=max(1, int(t.required_days or 1)) - 1))
        if actual_end < t.start_date:
            actual_end = t.start_date
        end_exclusive = datetime.combine(actual_end + timedelta(days=1), datetime.min.time()).replace(hour=18)
        uid = f"wnrh-{s.id}@scheduler"
        summary = f"{t.project_name}｜{t.site_city}｜{s.role}"
        desc = f"客户:{t.customer_name or ''}\n人数:{t.required_headcount} 天数:{t.required_days}\n负责人/成员:{a.name}"
        events.extend(
            [
                "BEGIN:VEVENT",
                f"UID:{uid}",
                f"DTSTAMP:{now}",
                f"DTSTART:{start.strftime('%Y%m%dT%H%M%S')}",
                f"DTEND:{end_exclusive.strftime('%Y%m%dT%H%M%S')}",
                f"SUMMARY:{ics_escape(summary)}",
                f"DESCRIPTION:{ics_escape(desc)}",
                "END:VEVENT",
            ]
        )
    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//WNRH Scheduler//CN",
        "CALSCALE:GREGORIAN",
        "X-WR-CALNAME:万宁睿和排班",
        *events,
        "END:VCALENDAR",
    ]
    return "\r\n".join(lines).encode("utf-8")


def update_auditor_record(
    auditor_id: int,
    name: str,
    gender: str,
    group_level: str,
    can_lead_team: bool,
    base_city: str,
    max_weekly_tasks: int,
    status_cn: str,
    monthly_cases: int,
    travel_days: int,
    continuous_days: int,
    last_task_end_city: str,
    last_task_end_date,
):
    with db_session() as db:
        obj = db.query(Auditor).filter(Auditor.id == int(auditor_id)).first()
        if not obj:
            st.error("未找到对应稽查员记录")
            return False

        obj.name = normalize_text(name)
        obj.gender = normalize_text(gender) or "女"
        obj.group_level = normalize_text(group_level) or "B"
        obj.can_lead_team = bool(can_lead_team)
        obj.base_city = normalize_text(base_city)
        obj.max_weekly_tasks = int(max_weekly_tasks or 0)
        obj.status = STATUS_MAP.get(normalize_text(status_cn), "active")
        obj.monthly_cases = int(monthly_cases or 0)
        obj.travel_days = int(travel_days or 0)
        obj.continuous_days = int(continuous_days or 0)
        obj.last_task_end_city = normalize_text(last_task_end_city) or None

        parsed_date = safe_parse_date(last_task_end_date)
        if parsed_date is not None:
            obj.last_task_end_date = parsed_date

        if not obj.name:
            st.error("姓名不能为空")
            db.rollback()
            return False
        if not obj.base_city:
            st.error("常驻城市不能为空")
            db.rollback()
            return False

        return safe_commit(db, f"更新稽查员#{auditor_id}")


def delete_auditor_record(auditor_id: int):
    with db_session() as db:
        obj = db.query(Auditor).filter(Auditor.id == int(auditor_id)).first()
        if not obj:
            st.error("未找到对应稽查员记录")
            return False
        db.query(Schedule).filter(Schedule.auditor_id == int(auditor_id)).delete()
        db.delete(obj)
        return safe_commit(db, f"删除稽查员#{auditor_id}")


def update_task_record(
    task_id: int,
    project_name: str,
    customer_name: str,
    need_expert: bool,
    required_headcount: int,
    required_days: int,
    required_gender: str,
    specified_auditors: str,
    preferred_experts: str,
    site_city: str,
    start_date_value,
    end_date_value,
):
    with db_session() as db:
        obj = db.query(Task).filter(Task.id == int(task_id)).first()
        if not obj:
            st.error("未找到对应任务记录")
            return False

        sd = safe_parse_date(start_date_value)
        ed = safe_parse_date(end_date_value)

        if sd is None:
            st.error("开始日期无效")
            return False

        if ed is None:
            ed = sd + timedelta(days=max(1, int(required_days or 1)) - 1)

        if ed < sd:
            ed = sd + timedelta(days=max(1, int(required_days or 1)) - 1)

        if ed < sd:
            st.error("结束日期不能早于开始日期")
            return False

        final_days = max(1, int(required_days or ((ed - sd).days + 1)))

        obj.project_name = normalize_text(project_name)
        obj.customer_name = normalize_text(customer_name) or None
        obj.need_expert = bool(need_expert)
        obj.required_headcount = max(1, int(required_headcount or 1))
        obj.required_days = final_days
        obj.required_gender = normalize_text(required_gender) or "不限"
        obj.specified_auditors = normalize_text(specified_auditors) or None
        obj.preferred_experts = normalize_text(preferred_experts) or None
        obj.site_city = normalize_text(site_city)
        obj.start_date = sd
        obj.end_date = ed

        if not obj.project_name:
            st.error("项目名称不能为空")
            db.rollback()
            return False
        if not obj.site_city:
            st.error("中心城市不能为空")
            db.rollback()
            return False

        schedules = db.query(Schedule).filter(Schedule.task_id == int(task_id)).all()
        for s in schedules:
            s.start_date = sd
            s.end_date = ed
            s.travel_to_city = obj.site_city

        return safe_commit(db, f"更新任务#{task_id}")


def delete_task_record(task_id: int):
    ensure_support_tables()
    with db_session() as db:
        obj = db.query(Task).filter(Task.id == int(task_id)).first()
        if not obj:
            st.error("未找到对应任务记录")
            return False
        db.query(Schedule).filter(Schedule.task_id == int(task_id)).delete()
        try:
            with engine.begin() as conn:
                conn.execute(text("DELETE FROM direct_assignments WHERE task_id = :task_id"), {"task_id": int(task_id)})
        except Exception:
            pass
        db.delete(obj)
        return safe_commit(db, f"删除任务#{task_id}")


def assign_team_to_task(db: Session, task: Task, leader_id: int, member_ids: list[int]):
    if db.query(Schedule).filter(Schedule.task_id == task.id).count() > 0:
        return False, "该任务已存在排班记录，不能重复排班"

    start_date = task.start_date
    end_date = task.end_date or (task.start_date + timedelta(days=max(1, int(task.required_days or 1)) - 1))
    selected_ids = [int(leader_id)] + [int(x) for x in member_ids if int(x) != int(leader_id)]

    for aid in selected_ids:
        existing = db.query(Schedule).filter(Schedule.auditor_id == aid).all()
        for s in existing:
            if not (end_date < s.start_date or s.end_date < start_date):
                return False, f"稽查员#{aid} 与已有任务时间冲突"

        auditor = db.query(Auditor).filter(Auditor.id == aid).first()
        if auditor and auditor.last_task_end_date and auditor.last_task_end_date >= start_date:
            return False, f"稽查员 {auditor.name} 的上次结束日期与本次开始日期冲突"

    def add_schedule(auditor_id: int, role: str):
        auditor = db.query(Auditor).filter(Auditor.id == auditor_id).first()
        if not auditor:
            return

        from_city = compute_from_city(auditor, task)
        km = get_distance_km(db, from_city, task.site_city)

        db.add(
            Schedule(
                task_id=task.id,
                auditor_id=auditor.id,
                role=role,
                start_date=start_date,
                end_date=end_date,
                travel_from_city=from_city,
                travel_to_city=task.site_city,
                distance_km=float(km),
                score=0.0,
                status="confirmed",
            )
        )

        auditor.monthly_cases = int(auditor.monthly_cases or 0) + 1
        days = (end_date - start_date).days + 1
        auditor.travel_days = int(auditor.travel_days or 0) + max(0, days)
        auditor.continuous_days = max(int(auditor.continuous_days or 0), days)
        auditor.last_task_end_city = task.site_city
        auditor.last_task_end_date = end_date

    add_schedule(int(leader_id), "leader")
    for mid in member_ids:
        if int(mid) != int(leader_id):
            add_schedule(int(mid), "member")

    return True, "ok"


def run_batch_schedule(db: Session, d1: date, d2: date, mode: str = "greedy"):
    if d2 < d1:
        d1, d2 = d2, d1

    scheduled_task_ids = {tid for (tid,) in db.query(Schedule.task_id).distinct().all()}
    tasks = db.query(Task).filter(Task.start_date >= d1, Task.start_date <= d2).all()
    tasks.sort(key=lambda t: (0 if t.need_expert else 1, -int(t.required_headcount or 1), t.start_date))

    auditors = db.query(Auditor).all()
    report = {"assigned": [], "skipped": [], "batch_week_counts": {}}

    for t in tasks:
        # 优先处理已定项目 / 硬指定项目
        direct_rows_existing = get_direct_assignments(int(t.id))
        if direct_rows_existing:
            ok, msg = sync_task_schedules_from_direct_assignments(t)
            if ok:
                report["assigned"].append({
                    "task_id": t.id,
                    "project": t.project_name,
                    "leader": direct_rows_existing[0].get("person_name", "") if direct_rows_existing else "",
                    "members": [r.get("person_name", "") for r in direct_rows_existing[1:]],
                })
            else:
                report["skipped"].append({"task_id": t.id, "project": t.project_name, "reason": msg})
            continue

        specified_names = parse_name_list(getattr(t, "specified_auditors", None))
        if specified_names:
            ok, msg = auto_fill_direct_assignments_from_specified(int(t.id), overwrite=True)
            if ok:
                report["assigned"].append({
                    "task_id": t.id,
                    "project": t.project_name,
                    "leader": specified_names[0] if specified_names else "",
                    "members": specified_names[1:],
                })
            else:
                report["skipped"].append({"task_id": t.id, "project": t.project_name, "reason": msg})
            continue

        if t.id in scheduled_task_ids:
            continue

        schedules_all = db.query(Schedule).all()
        candidates = build_candidates(db, t, auditors, schedules_all)
        team = propose_team(t, candidates)

        if mode == "optimized" and candidates:
            avg_cases = float(sum(int(a.monthly_cases or 0) for a in auditors) / max(1, len(auditors)))
            leader_pool = [c for c in candidates if c.can_lead_team]
            if t.need_expert:
                leader_pool = [c for c in leader_pool if c.group_level == "A"]
            leader_pool = leader_pool[:5]
            member_pool_all = candidates[:12]
            auditor_lookup = {a.id: a for a in auditors}
            best_team = None
            best_obj = None

            from app.scheduler import TeamProposal

            for leader in leader_pool:
                member_pool = [c for c in member_pool_all if c.auditor_id != leader.auditor_id]
                need_n = max(0, int(t.required_headcount or 1) - 1)

                if need_n == 0:
                    cand_team = TeamProposal(
                        leader=leader,
                        members=[],
                        team_score=leader.score,
                        notes="optimized-single",
                    )
                    obj = team_objective(cand_team, auditor_lookup, avg_cases, report["batch_week_counts"])
                    if best_obj is None or obj < best_obj:
                        best_obj, best_team = obj, cand_team
                    continue

                base_members = member_pool[:need_n]
                if len(base_members) < need_n:
                    continue

                cand_team = TeamProposal(
                    leader=leader,
                    members=base_members,
                    team_score=leader.score + sum(m.score for m in base_members) / max(1, len(base_members)),
                    notes="optimized",
                )
                obj = team_objective(cand_team, auditor_lookup, avg_cases, report["batch_week_counts"])
                if best_obj is None or obj < best_obj:
                    best_obj, best_team = obj, cand_team

            if best_team:
                team = best_team

        if not team:
            report["skipped"].append({"task_id": t.id, "project": t.project_name, "reason": "无可用团队"})
            continue

        leader_id = int(team.leader.auditor_id)
        member_ids = [int(m.auditor_id) for m in team.members]

        ok, msg = assign_team_to_task(db, t, leader_id, member_ids)
        if not ok:
            db.rollback()
            report["skipped"].append({"task_id": t.id, "project": t.project_name, "reason": msg})
            continue

        for aid in [leader_id] + member_ids:
            report["batch_week_counts"][aid] = int(report["batch_week_counts"].get(aid, 0)) + 1

        if not safe_commit(db, context=f"批量排班 commit：task#{t.id} {t.project_name}"):
            report["skipped"].append({"task_id": t.id, "project": t.project_name, "reason": "数据库写入失败"})
            continue

        report["assigned"].append(
            {
                "task_id": t.id,
                "project": t.project_name,
                "leader": team.leader.auditor_name,
                "members": [m.auditor_name for m in team.members],
            }
        )

    return report



def ensure_support_tables():
    with engine.begin() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS part_time_staff (
                id INTEGER PRIMARY KEY,
                name TEXT NOT NULL UNIQUE,
                base_city TEXT,
                note TEXT,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT,
                updated_at TEXT
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS direct_assignments (
                id INTEGER PRIMARY KEY,
                task_id INTEGER NOT NULL,
                auditor_id INTEGER,
                person_name TEXT,
                is_part_time INTEGER NOT NULL DEFAULT 0,
                role TEXT,
                start_date TEXT,
                end_date TEXT,
                notes TEXT,
                created_at TEXT,
                updated_at TEXT
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS progress_targets (
                id INTEGER PRIMARY KEY,
                period_type TEXT,
                year INTEGER,
                period_value INTEGER,
                target_projects INTEGER DEFAULT 0,
                target_staffing INTEGER DEFAULT 0,
                updated_at TEXT
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS auditor_month_targets (
                id INTEGER PRIMARY KEY,
                auditor_id INTEGER,
                monthly_target INTEGER DEFAULT 4,
                updated_at TEXT
            )
        """))

    def _add_col_if_missing(table_name: str, column_name: str, ddl_sql: str):
        try:
            with engine.begin() as conn:
                if IS_SQLITE:
                    cols = conn.execute(text(f"PRAGMA table_info({table_name})")).mappings().all()
                    existing = {str(c.get("name")) for c in cols}
                    if column_name not in existing:
                        conn.execute(text(ddl_sql))
                else:
                    conn.execute(text(ddl_sql))
        except Exception:
            pass

    _add_col_if_missing("part_time_staff", "updated_at", "ALTER TABLE part_time_staff ADD COLUMN updated_at TEXT")
    _add_col_if_missing("direct_assignments", "updated_at", "ALTER TABLE direct_assignments ADD COLUMN updated_at TEXT")
    _add_col_if_missing("progress_targets", "updated_at", "ALTER TABLE progress_targets ADD COLUMN updated_at TEXT")
    _add_col_if_missing("auditor_month_targets", "updated_at", "ALTER TABLE auditor_month_targets ADD COLUMN updated_at TEXT")

def ensure_extra_tables():
    ensure_support_tables()

# 启动初始化：放到依赖函数全部定义完成之后再执行，避免 NameError
initialize_app_once()
bootstrap_auth_users_if_needed()


def parse_name_list(raw):
    s = normalize_text(raw)
    if not s:
        return []
    for sep in ["，", "、", ";", "；", "/", "|"]:
        s = s.replace(sep, ",")
    return [x.strip() for x in s.split(",") if x.strip()]


def get_part_time_staff_rows(active_only: bool = False) -> list[dict]:
    ensure_support_tables()
    sql = "SELECT id, name, base_city, note, is_active, created_at, updated_at FROM part_time_staff"
    params = {}
    if active_only:
        sql += " WHERE is_active = 1"
    sql += " ORDER BY is_active DESC, name ASC"
    with engine.begin() as conn:
        rows = conn.execute(text(sql), params).mappings().all()
    return [dict(r) for r in rows]


def save_part_time_staff(name: str, base_city: str = "", note: str = "", is_active: bool = True):
    ensure_support_tables()
    name = normalize_text(name)
    if not name:
        return False, "兼职姓名不能为空"
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with engine.begin() as conn:
        existing = conn.execute(text("SELECT id FROM part_time_staff WHERE name=:name"), {"name": name}).mappings().first()
        if existing:
            conn.execute(text("""
                UPDATE part_time_staff
                SET base_city=:base_city, note=:note, is_active=:is_active, updated_at=:updated_at
                WHERE name=:name
            """), {"name": name, "base_city": normalize_text(base_city), "note": normalize_text(note), "is_active": 1 if is_active else 0, "updated_at": now})
            return True, "兼职人员已更新"
        next_id = (conn.execute(text("SELECT COALESCE(MAX(id), 0) FROM part_time_staff")).scalar() or 0) + 1
        conn.execute(text("""
            INSERT INTO part_time_staff (id, name, base_city, note, is_active, created_at, updated_at)
            VALUES (:id, :name, :base_city, :note, :is_active, :created_at, :updated_at)
        """), {"id": next_id, "name": name, "base_city": normalize_text(base_city), "note": normalize_text(note), "is_active": 1 if is_active else 0, "created_at": now, "updated_at": now})
    return True, "兼职人员已保存"


def delete_part_time_staff(row_id: int):
    ensure_support_tables()
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM part_time_staff WHERE id=:id"), {"id": int(row_id)})
    return True, "兼职人员已删除"


def get_direct_assignments(task_id: int) -> list[dict]:
    ensure_support_tables()
    sql_with_updated = """
        SELECT id, task_id, auditor_id, person_name, is_part_time, role, start_date, end_date, notes, created_at, updated_at
        FROM direct_assignments
        WHERE task_id=:task_id
        ORDER BY start_date ASC, id ASC
    """
    sql_legacy = """
        SELECT id, task_id, auditor_id, person_name, is_part_time, role, start_date, end_date, notes, created_at
        FROM direct_assignments
        WHERE task_id=:task_id
        ORDER BY start_date ASC, id ASC
    """
    try:
        with engine.begin() as conn:
            rows = conn.execute(text(sql_with_updated), {"task_id": int(task_id)}).mappings().all()
    except Exception:
        with engine.begin() as conn:
            rows = conn.execute(text(sql_legacy), {"task_id": int(task_id)}).mappings().all()
    out = []
    for r in rows:
        item = dict(r)
        item.setdefault("updated_at", None)
        out.append(item)
    return out


def replace_direct_assignments(task_id: int, rows: list[dict]):
    ensure_support_tables()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM direct_assignments WHERE task_id=:task_id"), {"task_id": int(task_id)})
        next_id = (conn.execute(text("SELECT COALESCE(MAX(id), 0) FROM direct_assignments")).scalar() or 0) + 1
        for i, r in enumerate(rows or []):
            conn.execute(text("""
                INSERT INTO direct_assignments (id, task_id, auditor_id, person_name, is_part_time, role, start_date, end_date, notes, created_at, updated_at)
                VALUES (:id, :task_id, :auditor_id, :person_name, :is_part_time, :role, :start_date, :end_date, :notes, :created_at, :updated_at)
            """), {
                "id": next_id + i,
                "task_id": int(task_id),
                "auditor_id": int(r.get("auditor_id")) if r.get("auditor_id") not in (None, "") else None,
                "person_name": normalize_text(r.get("person_name")),
                "is_part_time": 1 if bool(r.get("is_part_time")) else 0,
                "role": "leader" if str(r.get("role")) == "leader" else "member",
                "start_date": d2s(safe_parse_date(r.get("start_date"))),
                "end_date": d2s(safe_parse_date(r.get("end_date"))),
                "notes": normalize_text(r.get("notes")),
                "created_at": now,
                "updated_at": now,
            })
    return True


def auto_fill_direct_assignments_from_specified(task_id: int, overwrite: bool = True):
    with db_session() as db:
        task = db.query(Task).filter(Task.id == int(task_id)).first()
        if not task:
            return False, "任务不存在"
        names = parse_name_list(getattr(task, "specified_auditors", None))
        if not names:
            return False, "未填写硬指定人员"
        existing = get_direct_assignments(int(task_id))
        if existing and not overwrite:
            return True, "已存在已定项目人员"
        auditors = db.query(Auditor).all()
        auditor_name_to_id = {a.name: int(a.id) for a in auditors}
        rows = []
        for idx, nm in enumerate(names):
            rows.append({
                "auditor_id": auditor_name_to_id.get(nm),
                "person_name": nm,
                "is_part_time": auditor_name_to_id.get(nm) is None,
                "role": "leader" if idx == 0 else "member",
                "start_date": task.start_date,
                "end_date": task.end_date or task.start_date,
                "notes": "由硬指定人员自动生成",
            })
    replace_direct_assignments(int(task_id), rows)
    with db_session() as db:
        task = db.query(Task).filter(Task.id == int(task_id)).first()
        if task:
            return sync_task_schedules_from_direct_assignments(task)
    return True, "已根据硬指定人员生成直录排班"


def sync_task_schedules_from_direct_assignments(task):
    rows = get_direct_assignments(int(task.id))
    if not rows:
        return False, "未找到已定项目人员"
    with db_session() as db:
        task_obj = db.query(Task).filter(Task.id == int(task.id)).first()
        if not task_obj:
            return False, "任务不存在"
        db.query(Schedule).filter(Schedule.task_id == int(task.id)).delete()
        auditor_lookup = {a.id: a for a in db.query(Auditor).all()}
        for r in rows:
            auditor_id = r.get("auditor_id")
            if not auditor_id:
                continue
            auditor = auditor_lookup.get(int(auditor_id))
            if not auditor:
                continue
            sd = safe_parse_date(r.get("start_date")) or task_obj.start_date
            ed = safe_parse_date(r.get("end_date")) or task_obj.end_date or task_obj.start_date
            from_city = compute_from_city(auditor, task_obj)
            km = float(get_distance_km(db, from_city, task_obj.site_city))
            db.add(Schedule(
                task_id=int(task_obj.id),
                auditor_id=int(auditor.id),
                role="leader" if str(r.get("role")) == "leader" else "member",
                start_date=sd,
                end_date=ed,
                travel_from_city=from_city,
                travel_to_city=task_obj.site_city,
                distance_km=km,
                score=0.0,
                status="confirmed",
            ))
        if not safe_commit(db, context=f"同步已定项目到排班 task#{task_obj.id}"):
            return False, "同步到排班表失败"
    return True, "已定项目人员已同步到排班记录"


def get_period_date_range(period_type: str, year: int, period_value: int):
    if period_type == "monthly":
        start_d = date(year, period_value, 1)
        next_month = (start_d.replace(day=28) + timedelta(days=4)).replace(day=1)
        end_d = next_month - timedelta(days=1)
    elif period_type == "quarterly":
        start_month = 1 + (int(period_value) - 1) * 3
        start_d = date(year, start_month, 1)
        end_month_start = date(year + (1 if start_month + 3 > 12 else 0), ((start_month + 2) % 12) + 1, 1)
        next_month = (end_month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
        end_d = next_month - timedelta(days=1)
    else:
        start_d = date(year, 1, 1)
        end_d = date(year, 12, 31)
    return start_d, end_d


def date_ranges_overlap(a_start, a_end, b_start, b_end) -> bool:
    if not a_start or not b_start:
        return False
    a_end = a_end or a_start
    b_end = b_end or b_start
    return not (a_end < b_start or b_end < a_start)


def get_target_row(period_type: str, year: int, period_value: int) -> dict:
    ensure_support_tables()
    with engine.begin() as conn:
        row = conn.execute(text("""
            SELECT id, period_type, year, period_value, target_projects, target_staffing, updated_at
            FROM progress_targets
            WHERE period_type=:period_type AND year=:year AND period_value=:period_value
        """), {"period_type": period_type, "year": int(year), "period_value": int(period_value)}).mappings().first()
    return dict(row) if row else {"target_projects": 0, "target_staffing": 0}


def save_target_row(period_type: str, year: int, period_value: int, target_projects: int, target_staffing: int):
    ensure_support_tables()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with engine.begin() as conn:
        row = conn.execute(text("""
            SELECT id FROM progress_targets
            WHERE period_type=:period_type AND year=:year AND period_value=:period_value
        """), {"period_type": period_type, "year": int(year), "period_value": int(period_value)}).mappings().first()
        if row:
            conn.execute(text("""
                UPDATE progress_targets
                SET target_projects=:target_projects, target_staffing=:target_staffing, updated_at=:updated_at
                WHERE id=:id
            """), {"id": int(row["id"]), "target_projects": int(target_projects), "target_staffing": int(target_staffing), "updated_at": now})
        else:
            next_id = (conn.execute(text("SELECT COALESCE(MAX(id), 0) FROM progress_targets")).scalar() or 0) + 1
            conn.execute(text("""
                INSERT INTO progress_targets (id, period_type, year, period_value, target_projects, target_staffing, updated_at)
                VALUES (:id, :period_type, :year, :period_value, :target_projects, :target_staffing, :updated_at)
            """), {"id": next_id, "period_type": period_type, "year": int(year), "period_value": int(period_value), "target_projects": int(target_projects), "target_staffing": int(target_staffing), "updated_at": now})
    return True


def get_monthly_target_map():
    ensure_support_tables()
    with engine.begin() as conn:
        rows = conn.execute(text("SELECT auditor_id, monthly_target FROM auditor_month_targets")).mappings().all()
    return {int(r["auditor_id"]): int(r.get("monthly_target") or 6) for r in rows}


def save_monthly_target(auditor_id: int, monthly_target: int):
    ensure_support_tables()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with engine.begin() as conn:
        row = conn.execute(text("SELECT auditor_id FROM auditor_month_targets WHERE auditor_id=:auditor_id"), {"auditor_id": int(auditor_id)}).mappings().first()
        if row:
            conn.execute(text("UPDATE auditor_month_targets SET monthly_target=:monthly_target, updated_at=:updated_at WHERE auditor_id=:auditor_id"),
                         {"auditor_id": int(auditor_id), "monthly_target": int(monthly_target), "updated_at": now})
        else:
            conn.execute(text("INSERT INTO auditor_month_targets (auditor_id, monthly_target, updated_at) VALUES (:auditor_id, :monthly_target, :updated_at)"),
                         {"auditor_id": int(auditor_id), "monthly_target": int(monthly_target), "updated_at": now})
    return True


def get_progress_stats(period_type: str, year: int, period_value: int):
    start_d, end_d = get_period_date_range(period_type, year, period_value)
    today_d = date.today()
    with db_session() as db:
        tasks = db.query(Task).order_by(Task.start_date.asc()).all()
        schedules = db.query(Schedule).options(joinedload(Schedule.auditor), joinedload(Schedule.task)).all()

    in_period_tasks = []
    for t in tasks:
        t_start = getattr(t, "start_date", None)
        t_end = getattr(t, "end_date", None) or t_start
        if date_ranges_overlap(t_start, t_end, start_d, end_d):
            in_period_tasks.append(t)

    recorded_projects = len({int(t.id) for t in in_period_tasks})
    completed_projects = len({
        int(t.id) for t in in_period_tasks
        if ((getattr(t, "end_date", None) or getattr(t, "start_date", None))
            and (getattr(t, "end_date", None) or getattr(t, "start_date", None)) < today_d)
    })

    staffing_scheduled = 0
    staffing_completed = 0
    for s in schedules:
        s_start = getattr(s, "start_date", None)
        s_end = getattr(s, "end_date", None) or s_start
        if date_ranges_overlap(s_start, s_end, start_d, end_d):
            staffing_scheduled += 1
            if s_end and s_end < today_d:
                staffing_completed += 1

    detail_rows = []
    for t in in_period_tasks:
        detail_rows.append({
            "项目名称": t.project_name,
            "客户名称": t.customer_name or "",
            "城市": t.site_city,
            "开始日期": d2s(t.start_date),
            "结束日期": d2s(t.end_date),
            "状态": "已完成" if ((t.end_date or t.start_date) < today_d) else "进行中/未完成",
            "硬指定": t.specified_auditors or "",
        })

    return start_d, end_d, recorded_projects, completed_projects, staffing_scheduled, staffing_completed, detail_rows


def get_yearly_month_progress(year: int):
    rows = []
    for m in range(1, 13):
        start_d, end_d, recorded_projects, completed_projects, staffing_scheduled, staffing_completed, detail_rows = get_progress_stats("monthly", year, m)
        target = get_target_row("monthly", year, m)
        rows.append({
            "月份": f"{m}月",
            "计划项目完成": int(target.get("target_projects", 0) or 0),
            "实际项目完成": int(completed_projects),
            "已排项目数": int(recorded_projects),
            "计划人员院次": int(target.get("target_staffing", 0) or 0),
            "实际人员院次": int(staffing_completed),
            "已排人员院次": int(staffing_scheduled),
        })
    return pd.DataFrame(rows)


def get_auditor_completion_df(period_type: str, year: int, period_value: int):
    start_d, end_d = get_period_date_range(period_type, year, period_value)
    months_multiplier = 12 if period_type == "yearly" else (3 if period_type == "quarterly" else 1)
    monthly_target_map = get_monthly_target_map()
    with db_session() as db:
        auditors = db.query(Auditor).order_by(Auditor.name.asc()).all()
        schedules = db.query(Schedule).options(joinedload(Schedule.auditor)).all()
    actual_map = {}
    for s in schedules:
        s_start = getattr(s, "start_date", None)
        s_end = getattr(s, "end_date", None) or s_start
        if date_ranges_overlap(s_start, s_end, start_d, end_d):
            actual_map[int(s.auditor_id)] = int(actual_map.get(int(s.auditor_id), 0)) + 1
    rows = []
    for a in auditors:
        monthly_target = int(monthly_target_map.get(int(a.id), 6) or 6)
        plan = monthly_target * months_multiplier
        actual = int(actual_map.get(int(a.id), 0))
        rows.append({
            "稽查员": a.name,
            "月度计划院次": monthly_target,
            "本周期计划院次": plan,
            "本周期实际院次": actual,
            "完成率": round(actual / plan * 100.0, 1) if plan else 0.0,
        })
    return pd.DataFrame(rows)


def build_calendar_detail_map(merged_rows):
    event_map = {}
    for s in merged_rows:
        cur = s.get("start_date")
        end_d = s.get("end_date")
        while cur and end_d and cur <= end_d:
            event_map.setdefault(cur, []).append(s)
            cur += timedelta(days=1)
    return event_map


def render_day_detail_panel(calendar_event_map: dict):
    st.subheader("点击查看某一天的详细安排")
    day_options = sorted(calendar_event_map.keys())
    if not day_options:
        st.info("本月暂无安排")
        return
    selected_day = st.selectbox("选择日期", options=day_options, format_func=lambda d: d.strftime("%Y-%m-%d"), key="calendar_day_detail")
    render_calendar_day_popover(selected_day, calendar_event_map.get(selected_day, []), use_explicit_header=True)


def render_calendar_day_popover(day_obj: date, items: list[dict], use_explicit_header: bool = False):
    if use_explicit_header:
        st.markdown(f"**{day_obj.strftime('%Y-%m-%d')}**")
    if not items:
        st.info("当天暂无安排")
        return

    st.markdown(
        """
        <style>
        div[data-testid="stPopoverBody"]{
            min-width: 480px !important;
            max-width: 620px !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    grouped = {}
    for item in items:
        key = (
            item.get("project_name") or "",
            item.get("site_city") or "",
            d2s(item.get("start_date")),
            d2s(item.get("end_date")),
            "已定直录" if item.get("source") == "direct" else "标准排班",
        )
        grouped.setdefault(key, {"leaders": [], "members": []})
        person = item.get("auditor_name") or item.get("person_name") or ""
        if item.get("role") == "leader":
            if person and person not in grouped[key]["leaders"]:
                grouped[key]["leaders"].append(person)
        else:
            if person and person not in grouped[key]["members"]:
                grouped[key]["members"].append(person)

    rows = []
    for (project_name, site_city, start_s, end_s, source_label), people in grouped.items():
        merged_people = []
        merged_people.extend(people["leaders"])
        merged_people.extend([p for p in people["members"] if p not in merged_people])
        rows.append({
            "项目": project_name,
            "城市": site_city,
            "人员": "、".join(merged_people),
            "开始": start_s,
            "结束": end_s,
            "来源": source_label,
        })
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True, height=min(260, 80 + len(rows) * 34))


def load_day_marks():
    try:
        p = Path(__file__).resolve().parent / "app" / "holidays_cn.json"
        if p.exists():
            obj = json.loads(p.read_text(encoding="utf-8"))
            items = obj.get("items", [])
            if isinstance(items, list):
                return items
    except Exception:
        pass
    return []


# -------------------- 侧边栏 --------------------
st.sidebar.title(APP_NAME)
st.sidebar.caption(f"当前用户：{st.session_state.get('login_user', '')}")

if st.sidebar.button("退出登录", key="logout_btn"):
    st.session_state["logged_in"] = False
    st.session_state["is_admin"] = False
    st.session_state["is_super_admin"] = False
    st.session_state.pop("login_user", None)
    st.session_state["allowed_pages"] = DEFAULT_NORMAL_PAGES[:]
    clear_runtime_caches_after_data_change()
    st.rerun()

current_user = st.session_state.get("login_user", "")
is_admin = bool(st.session_state.get("is_admin", False))
allowed_pages = st.session_state.get("allowed_pages") or get_user_allowed_pages(current_user)
allowed_pages = _normalize_pages(allowed_pages) if not is_admin else ALL_PAGES[:]
st.session_state["allowed_pages"] = allowed_pages

page = st.sidebar.radio(
    label="",
    options=allowed_pages,
    key="nav_radio",
    label_visibility="collapsed",
)

st.sidebar.caption(f"当前位置：{page}")
st.title(APP_NAME)
st.caption(f"当前页面：{page}")

if (not is_admin) and (page not in allowed_pages):
    st.error("当前账号无权限访问该板块，请联系主管理员开通。")
    st.stop()

# -------------------- 智能排班 --------------------
if page == "智能排班":
    st.subheader("智能排班")
    st.caption("先按硬约束筛选，再按距离优先 + 适度负荷均衡评分推荐。")

    with db_session() as db:
        tasks = db.query(Task).order_by(Task.id.desc()).all()
        schedules_recent = (
            db.query(Schedule)
            .options(joinedload(Schedule.task), joinedload(Schedule.auditor))
            .order_by(Schedule.id.desc())
            .limit(120)
            .all()
        )
        schedules_recent_rows = []
        for s in schedules_recent:
            schedules_recent_rows.append(
                {
                    "ID": s.id,
                    "任务": f"#{s.task_id} {(s.task.project_name if s.task else '')}",
                    "人员": f"#{s.auditor_id} {(s.auditor.name if s.auditor else '')} ({(s.auditor.group_level if s.auditor else '')})",
                    "角色": s.role,
                    "时间": f"{d2s(s.start_date)} ~ {d2s(s.end_date)}",
                    "路线": f"{s.travel_from_city} → {s.travel_to_city}",
                    "km": round(float(s.distance_km or 0), 1),
                }
            )

    if not tasks:
        st.info("请先在【任务管理】中录入任务。")
    else:
        task_options = {
            f"#{t.id} {t.project_name}｜{t.site_city}｜{d2s(t.start_date)}｜{t.required_days}天｜{t.required_headcount}人": t.id
            for t in tasks
        }
        selected_label = st.selectbox("选择任务", list(task_options.keys()), key="smart_task_select")
        selected_task_id = task_options[selected_label]

        col_a, _ = st.columns([1, 3])
        if col_a.button("生成推荐", type="primary", key="gen_reco_btn"):
            with db_session() as db:
                task = db.query(Task).filter(Task.id == selected_task_id).first()
                auditors = db.query(Auditor).all()
                schedules_all = db.query(Schedule).all()
                candidates = build_candidates(db, task, auditors, schedules_all) if task else []
                team = propose_team(task, candidates) if task else None
                st.session_state["recommend_result"] = {
                    "task_id": selected_task_id,
                    "candidates": candidates[:25],
                    "team": team,
                    "error": None if team else "无可用团队方案",
                }
            st.rerun()

        rec = st.session_state.get("recommend_result")
        if rec and rec.get("task_id") == selected_task_id:
            with db_session() as db:
                task = db.query(Task).filter(Task.id == selected_task_id).first()
            if task:
                st.info(
                    f"已选择：{task.project_name}（{task.site_city}，{d2s(task.start_date)}，{task.required_days}天，{task.required_headcount}人；需要A带队：{'是' if task.need_expert else '否'}）"
                )
            if rec.get("error"):
                st.error(rec["error"])
            team = rec.get("team")
            if team:
                st.subheader("系统推荐团队方案")
                st.write(
                    f"**负责人：** {team.leader.auditor_name}（{team.leader.group_level}，{'可带队' if team.leader.can_lead_team else '不可带队'}，出发地 {team.leader.from_city}，{team.leader.km:.0f}km，评分 {team.leader.score}）"
                )
                if team.members:
                    st.write(
                        "**组员：** "
                        + "； ".join(
                            [
                                f"{m.auditor_name}（{m.group_level}，{m.from_city}，{m.km:.0f}km，评分 {m.score}）"
                                for m in team.members
                            ]
                        )
                    )
                else:
                    st.write("**组员：** 无")
                st.caption(f"{team.notes}｜团队评分 {team.team_score}")
                default_member_ids = ",".join([str(m.auditor_id) for m in team.members])
                member_ids_text = st.text_input("确认指派前，可手工调整组员ID（逗号分隔）", value=default_member_ids, key="member_ids_text")
                if st.button("确认指派", type="primary", key="confirm_assign_btn"):
                    ids = [x for x in re.split(r"[，,\s]+", member_ids_text.strip()) if x.strip()]
                    member_ids = []
                    for x in ids:
                        try:
                            member_ids.append(int(x))
                        except Exception:
                            pass
                    with db_session() as db:
                        task = db.query(Task).filter(Task.id == selected_task_id).first()
                        ok, msg = assign_team_to_task(db, task, int(team.leader.auditor_id), member_ids)
                        if not ok:
                            db.rollback()
                            st.error(msg)
                            st.stop()
                        if not safe_commit(db, context=f"确认指派：task#{selected_task_id}"):
                            st.stop()
                    clear_runtime_caches_after_data_change()
                    st.success("已确认指派")
                    st.rerun()

            cands = rec.get("candidates") or []
            if cands:
                st.subheader("候选人 TOP25")
                rows = []
                for i, c in enumerate(cands, start=1):
                    rows.append(
                        {
                            "排名": i,
                            "姓名": c.auditor_name,
                            "组别": c.group_level,
                            "带队": "是" if c.can_lead_team else "否",
                            "出发地": c.from_city,
                            "距离(km)": round(float(c.km), 1),
                            "评分": c.score,
                            "解释": c.explain,
                        }
                    )
                show_table(rows, 420)

    st.subheader("最近排班记录（TOP120）")
    show_table(schedules_recent_rows, 360)
    if schedules_recent_rows:
        delete_sid = st.selectbox("删除排班记录（按ID）", [r["ID"] for r in schedules_recent_rows], key="delete_schedule_select")
        if st.button("删除所选排班记录", key="delete_schedule_btn"):
            with db_session() as db:
                obj = db.query(Schedule).filter(Schedule.id == delete_sid).first()
                if obj:
                    db.delete(obj)
                    if not safe_commit(db, context=f"删除排班记录：schedule#{delete_sid}"):
                        st.stop()
            clear_runtime_caches_after_data_change()
            st.success("已删除")
            st.rerun()

# -------------------- 批量排班 --------------------
elif page == "批量排班":
    st.subheader("批量排班")
    st.caption("只会处理未排过的任务；按 need_expert 优先 > 人数多优先 > 开始日期早 排序。")

    c1, c2, c3 = st.columns([1, 1, 1])
    date_start = c1.date_input("开始日期", value=date.today(), key="batch_start")
    date_end = c2.date_input("结束日期", value=date.today() + timedelta(days=30), key="batch_end")
    mode = c3.selectbox(
        "模式",
        ["greedy", "optimized"],
        format_func=lambda x: "快速模式（优先效率）" if x == "greedy" else "优化模式（优先成本与均衡）",
        key="batch_mode",
    )
    if st.button("开始批量排班", type="primary", key="batch_run_btn"):
        with db_session() as db:
            report = run_batch_schedule(db, date_start, date_end, mode)
        st.session_state["batch_report"] = report
        st.rerun()

    report = st.session_state.get("batch_report")
    if report:
        c1, c2 = st.columns(2)
        with c1:
            st.subheader(f"已自动排班（{len(report.get('assigned', []))}）")
            if report.get("assigned"):
                for a in report["assigned"]:
                    st.write(f"**#{a['task_id']} {a['project']}**")
                    st.caption(f"负责人：{a['leader']}；组员：{', '.join(a['members']) if a['members'] else '无'}")
            else:
                st.info("无")
        with c2:
            st.subheader(f"跳过任务（{len(report.get('skipped', []))}）")
            if report.get("skipped"):
                for s in report["skipped"]:
                    st.write(f"**#{s['task_id']} {s['project']}**")
                    st.caption(f"原因：{s['reason']}")
            else:
                st.info("无")

# -------------------- 稽查员管理 --------------------
elif page == "稽查员管理":
    st.subheader("稽查员管理")

    with st.form("auditor_form", clear_on_submit=True):
        c1, c2, c3, c4 = st.columns(4)
        name = c1.text_input("姓名*")
        gender = c2.selectbox("性别*", ["女", "男"], index=0)
        group_level = c3.selectbox("等级*", ["A", "B", "C"], index=1)
        can_lead = c4.selectbox("可带队*", ["是", "否"], index=0)

        c5, c6, c7, c8 = st.columns(4)
        base_city = c5.text_input("常驻城市*")
        max_weekly_tasks = c6.number_input("每周上限", min_value=0, value=1, step=1)
        status_cn = c7.selectbox("状态", ["在岗", "请假", "冻结"])
        monthly_cases = c8.number_input("本月已排院次", min_value=0, value=0, step=1)

        c9, c10, c11, c12 = st.columns(4)
        travel_days = c9.number_input("本月差旅天数", min_value=0, value=0, step=1)
        continuous_days = c10.number_input("连续工作天数", min_value=0, value=0, step=1)
        last_city = c11.text_input("上次结束城市（可空）")
        last_date = c12.date_input("上次结束日期*", value=date.today())

        if st.form_submit_button("新增稽查员", type="primary"):
            if not name.strip() or not base_city.strip():
                st.error("姓名、常驻城市必填。")
            else:
                with db_session() as db:
                    db.add(
                        Auditor(
                            name=name.strip(),
                            gender=gender,
                            group_level=group_level,
                            can_lead_team=(can_lead == "是"),
                            base_city=base_city.strip(),
                            max_weekly_tasks=int(max_weekly_tasks),
                            status=STATUS_MAP[status_cn],
                            monthly_cases=int(monthly_cases),
                            travel_days=int(travel_days),
                            continuous_days=int(continuous_days),
                            last_task_end_city=last_city.strip() or None,
                            last_task_end_date=last_date,
                        )
                    )
                    if not safe_commit(db, context=f"新增稽查员：{name.strip()}"):
                        st.stop()
                with db_session() as db:
                    task_obj = db.query(Task).order_by(Task.id.desc()).first()
                if task_obj and parse_name_list(task_obj.specified_auditors or ""):
                    auto_fill_direct_assignments_from_specified(int(task_obj.id), overwrite=True)
                clear_runtime_caches_after_data_change()
                st.success("已新增")
                st.rerun()

    with db_session() as db:
        auditors = db.query(Auditor).order_by(Auditor.id.desc()).all()

    rows = []
    for a in auditors:
        rows.append(
            {
                "ID": a.id,
                "姓名": a.name,
                "性别": a.gender,
                "等级": a.group_level,
                "可带队": "是" if a.can_lead_team else "否",
                "常驻城市": a.base_city,
                "周上限": a.max_weekly_tasks,
                "状态": STATUS_MAP_REV.get(a.status, a.status),
                "本月院次": a.monthly_cases,
                "差旅天数": a.travel_days,
                "连续天数": a.continuous_days,
                "上次结束城市": a.last_task_end_city or "",
                "上次结束日期": d2s(a.last_task_end_date),
            }
        )

    show_table(rows, 320)

    if auditors:
        auditor_options = {
            f"#{a.id} {a.name}｜{a.base_city}｜{a.group_level}": a.id
            for a in auditors
        }
        selected_auditor_label = st.selectbox("选择要编辑的稽查员", list(auditor_options.keys()), key="edit_auditor_select")
        selected_auditor_id = auditor_options[selected_auditor_label]
        selected_auditor = next((a for a in auditors if a.id == selected_auditor_id), None)

        if selected_auditor:
            st.divider()
            st.subheader(f"编辑稽查员 #{selected_auditor.id}")

            with st.form("edit_auditor_form", clear_on_submit=False):
                c1, c2, c3, c4 = st.columns(4)
                edit_name = c1.text_input("姓名*", value=selected_auditor.name or "")
                edit_gender = c2.selectbox("性别*", ["女", "男"], index=0 if (selected_auditor.gender or "女") == "女" else 1)
                edit_group = c3.selectbox("等级*", ["A", "B", "C"], index=["A", "B", "C"].index(selected_auditor.group_level or "B"))
                edit_can_lead = c4.selectbox("可带队*", ["是", "否"], index=0 if selected_auditor.can_lead_team else 1)

                c5, c6, c7, c8 = st.columns(4)
                edit_base_city = c5.text_input("常驻城市*", value=selected_auditor.base_city or "")
                edit_max_weekly_tasks = c6.number_input("每周上限", min_value=0, value=int(selected_auditor.max_weekly_tasks or 0), step=1)
                edit_status = c7.selectbox(
                    "状态",
                    ["在岗", "请假", "冻结"],
                    index=["在岗", "请假", "冻结"].index(STATUS_MAP_REV.get(selected_auditor.status, "在岗")),
                )
                edit_monthly_cases = c8.number_input("本月已排院次", min_value=0, value=int(selected_auditor.monthly_cases or 0), step=1)

                c9, c10, c11, c12 = st.columns(4)
                edit_travel_days = c9.number_input("本月差旅天数", min_value=0, value=int(selected_auditor.travel_days or 0), step=1)
                edit_continuous_days = c10.number_input("连续工作天数", min_value=0, value=int(selected_auditor.continuous_days or 0), step=1)
                edit_last_city = c11.text_input("上次结束城市（可空）", value=selected_auditor.last_task_end_city or "")
                edit_last_date = c12.date_input("上次结束日期*", value=selected_auditor.last_task_end_date or date.today())

                b1, b2 = st.columns(2)
                save_edit = b1.form_submit_button("保存当前稽查员修改", type="primary")
                delete_edit = b2.form_submit_button("删除当前稽查员")

            if save_edit:
                ok = update_auditor_record(
                    auditor_id=selected_auditor.id,
                    name=edit_name,
                    gender=edit_gender,
                    group_level=edit_group,
                    can_lead_team=(edit_can_lead == "是"),
                    base_city=edit_base_city,
                    max_weekly_tasks=int(edit_max_weekly_tasks),
                    status_cn=edit_status,
                    monthly_cases=int(edit_monthly_cases),
                    travel_days=int(edit_travel_days),
                    continuous_days=int(edit_continuous_days),
                    last_task_end_city=edit_last_city,
                    last_task_end_date=edit_last_date,
                )
                if ok:
                    clear_runtime_caches_after_data_change()
                    st.success("稽查员修改已保存")
                    st.rerun()

            if delete_edit:
                ok = delete_auditor_record(selected_auditor.id)
                if ok:
                    clear_runtime_caches_after_data_change()
                    st.success("稽查员已删除")
                    st.rerun()

# -------------------- 任务管理 --------------------
elif page == "任务管理":
    st.subheader("任务管理")

    with st.form("task_form", clear_on_submit=True):
        c1, c2, c3 = st.columns(3)
        project_name = c1.text_input("项目名称*")
        customer_name = c2.text_input("客户/申办方（可空）")
        need_expert = c3.selectbox("需要A带队", ["否", "是"])

        c4, c5, c6, c7 = st.columns(4)
        required_headcount = c4.number_input("所需人数", min_value=1, value=1, step=1)
        required_days = c5.number_input("任务天数", min_value=1, value=1, step=1)
        required_gender = c6.selectbox("性别要求", ["不限", "男", "女"])
        site_city = c7.text_input("中心城市*")

        c8, c9, c10 = st.columns(3)
        specified = c8.text_input("硬指定人员（可空，支持 ，、分隔）")
        preferred = c9.text_input("软指定专家/老师（可空）")
        start_date = c10.date_input("开始日期*", value=date.today())
        default_end = start_date + timedelta(days=max(1, int(required_days)) - 1)
        end_date = st.date_input("结束日期*", value=default_end)

        if st.form_submit_button("新增任务", type="primary"):
            if not project_name.strip() or not site_city.strip():
                st.error("项目名称、中心城市必填。")
            elif end_date < start_date:
                st.error("结束日期不能早于开始日期。")
            else:
                with db_session() as db:
                    db.add(
                        Task(
                            project_name=project_name.strip(),
                            customer_name=customer_name.strip() or None,
                            need_expert=(need_expert == "是"),
                            required_headcount=int(required_headcount),
                            required_days=int(required_days),
                            required_gender=required_gender,
                            specified_auditors=specified.strip() or None,
                            preferred_experts=preferred.strip() or None,
                            site_city=site_city.strip(),
                            start_date=start_date,
                            end_date=end_date,
                        )
                    )
                    if not safe_commit(db, context=f"新增任务：{project_name.strip()}"):
                        st.stop()
                clear_runtime_caches_after_data_change()
                st.success("已新增")
                st.rerun()

    with db_session() as db:
        tasks = db.query(Task).order_by(Task.id.desc()).all()
        auditors = db.query(Auditor).order_by(Auditor.name.asc()).all()

    rows = []
    for t in tasks:
        rows.append(
            {
                "ID": t.id,
                "项目": t.project_name,
                "客户": t.customer_name or "",
                "需要A": "是" if t.need_expert else "否",
                "人数": t.required_headcount,
                "天数": t.required_days,
                "性别": t.required_gender,
                "硬指定": t.specified_auditors or "",
                "软指定": t.preferred_experts or "",
                "城市": t.site_city,
                "开始": d2s(t.start_date),
                "结束": d2s(t.end_date),
            }
        )

    show_table(rows, 320)

    if tasks:
        task_options = {
            f"#{t.id} {t.project_name}｜{t.site_city}｜{d2s(t.start_date)}": t.id
            for t in tasks
        }
        selected_task_label = st.selectbox("选择要编辑的任务", list(task_options.keys()), key="edit_task_select")
        selected_task_id = task_options[selected_task_label]
        selected_task = next((t for t in tasks if t.id == selected_task_id), None)

        if selected_task:
            st.divider()
            st.subheader(f"编辑任务 #{selected_task.id}")

            with st.form("edit_task_form", clear_on_submit=False):
                c1, c2, c3 = st.columns(3)
                edit_project_name = c1.text_input("项目名称*", value=selected_task.project_name or "")
                edit_customer_name = c2.text_input("客户/申办方（可空）", value=selected_task.customer_name or "")
                edit_need_expert = c3.selectbox("需要A带队", ["否", "是"], index=1 if selected_task.need_expert else 0)

                c4, c5, c6, c7 = st.columns(4)
                edit_required_headcount = c4.number_input("所需人数", min_value=1, value=int(selected_task.required_headcount or 1), step=1)
                edit_required_days = c5.number_input("任务天数", min_value=1, value=int(selected_task.required_days or 1), step=1)
                edit_required_gender = c6.selectbox(
                    "性别要求",
                    ["不限", "男", "女"],
                    index=["不限", "男", "女"].index(selected_task.required_gender or "不限"),
                )
                edit_site_city = c7.text_input("中心城市*", value=selected_task.site_city or "")

                c8, c9, c10 = st.columns(3)
                edit_specified = c8.text_input("硬指定人员（可空）", value=selected_task.specified_auditors or "")
                edit_preferred = c9.text_input("软指定专家/老师（可空）", value=selected_task.preferred_experts or "")
                edit_start_date = c10.date_input("开始日期*", value=selected_task.start_date or date.today())
                edit_end_date = st.date_input("结束日期*", value=selected_task.end_date or edit_start_date)

                b1, b2 = st.columns(2)
                save_task = b1.form_submit_button("保存当前任务修改", type="primary")
                delete_task = b2.form_submit_button("删除当前任务")

            if save_task:
                ok = update_task_record(
                    task_id=selected_task.id,
                    project_name=edit_project_name,
                    customer_name=edit_customer_name,
                    need_expert=(edit_need_expert == "是"),
                    required_headcount=int(edit_required_headcount),
                    required_days=int(edit_required_days),
                    required_gender=edit_required_gender,
                    specified_auditors=edit_specified,
                    preferred_experts=edit_preferred,
                    site_city=edit_site_city,
                    start_date_value=edit_start_date,
                    end_date_value=edit_end_date,
                )
                if ok:
                    clear_runtime_caches_after_data_change()
                    st.success("任务修改已保存")
                    st.rerun()

            if delete_task:
                ok = delete_task_record(selected_task.id)
                if ok:
                    clear_runtime_caches_after_data_change()
                    st.success("任务已删除")
                    st.rerun()

            st.divider()
            st.subheader("已定项目人员录入 / 直录排班")
            st.caption("适用于项目经理已完成排班，不需要系统推荐。支持内部稽查员与兼职人员，且可分别设置每个人在项目中的起止日期。")

            part_time_rows = get_part_time_staff_rows(active_only=True)
            auditor_name_to_id = {a.name: a.id for a in auditors}
            auditor_name_options = [a.name for a in auditors]
            part_time_options = [r["name"] for r in part_time_rows]

            existing_direct = get_direct_assignments(int(selected_task.id))
            if existing_direct:
                direct_df = pd.DataFrame([
                    {
                        "类型": "兼职" if bool(r.get("is_part_time")) else "内部稽查员",
                        "人员姓名": r.get("person_name", ""),
                        "角色": "组长" if str(r.get("role", "")) == "leader" else "成员",
                        "开始日期": str(r.get("start_date", "")),
                        "结束日期": str(r.get("end_date", "")),
                        "备注": r.get("notes", "") or "",
                    }
                    for r in existing_direct
                ])
            else:
                default_names = parse_name_list(selected_task.specified_auditors or "")
                if default_names:
                    direct_df = pd.DataFrame([
                        {
                            "类型": "内部稽查员" if nm in auditor_name_to_id else "兼职",
                            "人员姓名": nm,
                            "角色": "成员",
                            "开始日期": d2s(selected_task.start_date),
                            "结束日期": d2s(selected_task.end_date),
                            "备注": "",
                        }
                        for nm in default_names
                    ])
                else:
                    direct_df = pd.DataFrame(columns=["类型", "人员姓名", "角色", "开始日期", "结束日期", "备注"])

            with st.form("direct_assign_form", clear_on_submit=False):
                edited_direct = st.data_editor(
                    direct_df,
                    use_container_width=True,
                    hide_index=True,
                    num_rows="dynamic",
                    key=f"direct_assign_editor_{selected_task.id}",
                    column_config={
                        "类型": st.column_config.SelectboxColumn(options=["内部稽查员", "兼职"]),
                        "人员姓名": st.column_config.TextColumn(help="内部稽查员可填写系统内姓名；兼职可填写兼职库姓名"),
                        "角色": st.column_config.SelectboxColumn(options=["组长", "成员"]),
                        "开始日期": st.column_config.TextColumn(help="格式 YYYY-MM-DD"),
                        "结束日期": st.column_config.TextColumn(help="格式 YYYY-MM-DD"),
                        "备注": st.column_config.TextColumn(),
                    },
                )
                c1, c2 = st.columns(2)
                save_direct = c1.form_submit_button("保存已定项目人员")
                sync_direct = c2.form_submit_button("按已定人员直接录入排班", type="primary")

            def _normalize_direct_rows(df_in):
                rows_out = []
                for _, r in pd.DataFrame(df_in).iterrows():
                    person_name = str(r.get("人员姓名", "")).strip()
                    if not person_name:
                        continue
                    role = "leader" if str(r.get("角色", "")).strip() == "组长" else "member"
                    is_part_time = str(r.get("类型", "")).strip() == "兼职"
                    sd = safe_parse_date(r.get("开始日期"))
                    ed = safe_parse_date(r.get("结束日期"))
                    if not sd or not ed:
                        continue
                    rows_out.append(
                        {
                            "auditor_id": None if is_part_time else auditor_name_to_id.get(person_name),
                            "person_name": person_name,
                            "is_part_time": is_part_time,
                            "role": role,
                            "start_date": sd,
                            "end_date": ed,
                            "notes": str(r.get("备注", "")).strip(),
                        }
                    )
                return rows_out

            if save_direct:
                rows_to_save = _normalize_direct_rows(edited_direct)
                replace_direct_assignments(int(selected_task.id), rows_to_save)
                ok, msg = sync_task_schedules_from_direct_assignments(selected_task)
                clear_runtime_caches_after_data_change()
                if ok:
                    st.success("已定项目人员已保存，并同步到排班记录")
                else:
                    st.warning(msg)
                st.rerun()

            if sync_direct:
                rows_to_save = _normalize_direct_rows(edited_direct)
                if not rows_to_save:
                    st.error("请先录入至少1条已定项目人员")
                else:
                    replace_direct_assignments(int(selected_task.id), rows_to_save)
                    ok, msg = sync_task_schedules_from_direct_assignments(selected_task)
                    if ok:
                        clear_runtime_caches_after_data_change()
                        st.success(msg)
                        st.rerun()
                    else:
                        st.error(msg)


# -------------------- 页面：指标统计 --------------------
elif page == "指标统计":
    st.subheader("指标统计")
    st.caption("支持录入院次指标数量，并自动统计目标、已排、实际，以及年度月度趋势和每位稽查员月/季/年的计划与实际完成情况。")

    c1, c2, c3 = st.columns(3)
    period_type = c1.selectbox("统计周期", ["monthly", "quarterly", "yearly"], format_func=lambda x: {"monthly":"月度","quarterly":"季度","yearly":"年度"}[x])
    year = c2.number_input("年份", min_value=2024, max_value=2035, value=date.today().year, step=1)
    if period_type == "monthly":
        period_value = c3.selectbox("月份", list(range(1, 13)), index=max(0, date.today().month - 1))
    elif period_type == "quarterly":
        period_value = c3.selectbox("季度", [1, 2, 3, 4], index=(date.today().month - 1)//3)
    else:
        period_value = 0
        c3.markdown("**全年**")

    target = get_target_row(period_type, int(year), int(period_value))
    with st.form("target_form", clear_on_submit=False):
        c1, c2 = st.columns(2)
        target_projects = c1.number_input("院次指标数量", min_value=0, value=int(target.get("target_projects", 0) or 0), step=1)
        target_staffing = c2.number_input("人员院次安排数量", min_value=0, value=int(target.get("target_staffing", 0) or 0), step=1)
        if st.form_submit_button("保存指标", type="primary"):
            save_target_row(period_type, int(year), int(period_value), int(target_projects), int(target_staffing))
            st.success("指标已保存")
            st.rerun()

    start_d, end_d, scheduled_projects, actual_projects, scheduled_staffing, actual_staffing, detail_rows = get_progress_stats(period_type, int(year), int(period_value))
    target = get_target_row(period_type, int(year), int(period_value))
    t_projects = int(target.get("target_projects", 0) or 0)
    t_staff = int(target.get("target_staffing", 0) or 0)

    summary_df = pd.DataFrame([
        {"指标": "项目院次数", "目标": t_projects, "已排": scheduled_projects, "实际": actual_projects, "完成率": f"{round(actual_projects / t_projects * 100, 1) if t_projects else 0.0}%"},
        {"指标": "人员院次安排数", "目标": t_staff, "已排": scheduled_staffing, "实际": actual_staffing, "完成率": f"{round(actual_staffing / t_staff * 100, 1) if t_staff else 0.0}%"},
    ])
    st.write(f"统计区间：{d2s(start_d)} ~ {d2s(end_d)}")
    st.dataframe(summary_df, use_container_width=True, hide_index=True)
    st.bar_chart(summary_df.set_index("指标")[["目标", "已排", "实际"]])

    st.subheader("项目完成进度及人员安排进度明细")
    if detail_rows:
        st.dataframe(pd.DataFrame(detail_rows), use_container_width=True, hide_index=True)
    else:
        st.info("该统计区间暂无项目数据")

    if period_type == "yearly":
        st.divider()
        st.subheader("年度视图：每月计划与实际完成")
        ym_df = get_yearly_month_progress(int(year))
        st.dataframe(ym_df, use_container_width=True, hide_index=True)
        st.bar_chart(ym_df.set_index("月份")[["计划项目完成", "实际项目完成", "已排项目数"]])
        st.line_chart(ym_df.set_index("月份")[["计划项目完成", "实际项目完成", "已排项目数"]])

    st.divider()
    st.subheader("稽查员月/季/年完成情况")
    with db_session() as db:
        auditors = db.query(Auditor).order_by(Auditor.name.asc()).all()
    target_map = get_monthly_target_map()
    if auditors:
        edit_rows = []
        for a in auditors:
            edit_rows.append({"稽查员": a.name, "每月计划院次": int(target_map.get(int(a.id), 6) or 6)})
        with st.form("auditor_month_target_form", clear_on_submit=False):
            edited_plan_df = st.data_editor(pd.DataFrame(edit_rows), use_container_width=True, hide_index=True, num_rows="fixed")
            if st.form_submit_button("保存每人每月计划院次", type="primary"):
                name_to_id = {a.name: int(a.id) for a in auditors}
                for _, r in pd.DataFrame(edited_plan_df).iterrows():
                    nm = normalize_text(r.get("稽查员"))
                    if nm in name_to_id:
                        save_monthly_target(name_to_id[nm], _safe_int(r.get("每月计划院次"), 6) or 6)
                st.success("已保存")
                st.rerun()

        auditor_df = get_auditor_completion_df(period_type, int(year), int(period_value))
        st.dataframe(auditor_df, use_container_width=True, hide_index=True)
        if not auditor_df.empty:
            st.bar_chart(auditor_df.set_index("稽查员")[["本周期计划院次", "本周期实际院次"]])
            st.line_chart(auditor_df.set_index("稽查员")[["本周期计划院次", "本周期实际院次"]])
# -------------------- 页面：兼职库 --------------------
elif page == "兼职库":
    st.subheader("兼职库")
    st.caption("兼职人员不会进入系统自动推荐，仅用于已定项目人员录入与直录排班场景。")

    with st.form("part_time_form", clear_on_submit=True):
        c1, c2, c3 = st.columns(3)
        pt_name = c1.text_input("兼职姓名*")
        pt_city = c2.text_input("常驻城市（可空）")
        pt_active = c3.selectbox("状态", ["启用", "停用"], index=0)
        pt_note = st.text_input("备注（可空）")
        if st.form_submit_button("保存兼职人员", type="primary"):
            ok, msg = save_part_time_staff(pt_name, pt_city, pt_note, is_active=(pt_active == "启用"))
            if ok:
                st.success(msg)
                st.rerun()
            else:
                st.error(msg)

    rows = get_part_time_staff_rows(active_only=False)
    if rows:
        show_table([
            {
                "ID": r["id"],
                "姓名": r["name"],
                "常驻城市": r.get("base_city") or "",
                "状态": "启用" if int(r.get("is_active", 0)) == 1 else "停用",
                "备注": r.get("note") or "",
            } for r in rows
        ], 280)

        deletable = {f"#{r['id']} {r['name']}": r["id"] for r in rows}
        c1, c2 = st.columns([2, 1])
        del_label = c1.selectbox("选择要删除的兼职人员", list(deletable.keys()))
        if c2.button("删除兼职人员"):
            ok, msg = delete_part_time_staff(deletable[del_label])
            if ok:
                st.success(msg)
                st.rerun()
            else:
                st.error(msg)
    else:
        st.info("暂无兼职人员")
# -------------------- 城市距离 --------------------
elif page == "城市距离":
    st.subheader("城市距离")
    st.caption("系统会优先读取距离表；若未命中，会尝试按城市坐标自动计算并写回缓存。")
    with st.form("distance_form", clear_on_submit=True):
        c1, c2, c3 = st.columns(3)
        from_city = c1.text_input("出发城市*")
        to_city = c2.text_input("到达城市*")
        km = c3.number_input("公里数", min_value=0.0, value=0.0, step=1.0)
        if st.form_submit_button("新增 / 更新", type="primary"):
            if not from_city.strip() or not to_city.strip():
                st.error("出发城市、到达城市必填。")
            else:
                with db_session() as db:
                    a = from_city.strip()
                    b = to_city.strip()
                    rec = db.query(CityDistance).filter(CityDistance.from_city == a, CityDistance.to_city == b).first()
                    if rec:
                        rec.km = float(km)
                    else:
                        db.add(CityDistance(from_city=a, to_city=b, km=float(km)))
                    if not safe_commit(db, context=f"城市距离新增/更新：{a}->{b}"):
                        st.stop()
                clear_runtime_caches_after_data_change()
                st.success("已保存")
                st.rerun()

    with db_session() as db:
        dists = db.query(CityDistance).order_by(CityDistance.id.desc()).limit(300).all()
    rows = [{"ID": d.id, "from": d.from_city, "to": d.to_city, "km": round(float(d.km or 0), 1)} for d in dists]
    show_table(rows)
    if dists:
        delete_id = st.selectbox("删除距离记录（按ID）", [d.id for d in dists], key="delete_dist_select")
        if st.button("删除所选距离记录", key="delete_dist_btn"):
            with db_session() as db:
                obj = db.query(CityDistance).filter(CityDistance.id == delete_id).first()
                if obj:
                    db.delete(obj)
                    if not safe_commit(db, context=f"删除城市距离：dist#{delete_id}"):
                        st.stop()
            clear_runtime_caches_after_data_change()
            st.success("已删除")
            st.rerun()

# -------------------- 城市坐标 --------------------
elif page == "城市坐标":
    st.subheader("城市坐标")
    st.caption("用于自动计算全国城市直线距离；CSV 格式：name,lat,lon。")
    with st.form("city_form", clear_on_submit=True):
        c1, c2, c3 = st.columns(3)
        name = c1.text_input("城市名*")
        lat = c2.number_input("纬度 lat", value=0.0, step=0.000001, format="%.6f")
        lon = c3.number_input("经度 lon", value=0.0, step=0.000001, format="%.6f")
        if st.form_submit_button("新增 / 更新", type="primary"):
            if not name.strip():
                st.error("城市名必填。")
            else:
                with db_session() as db:
                    nm = name.strip()
                    rec = db.query(City).filter(City.name == nm).first()
                    if rec:
                        rec.lat = float(lat)
                        rec.lon = float(lon)
                    else:
                        db.add(City(name=nm, lat=float(lat), lon=float(lon)))
                    if not safe_commit(db, context=f"城市坐标新增/更新：{nm}"):
                        st.stop()
                clear_runtime_caches_after_data_change()
                st.success("已保存")
                st.rerun()

    csv_file = st.file_uploader("批量导入城市坐标 CSV", type=["csv"], key="city_csv")
    if st.button("执行 CSV 导入", key="city_csv_import_btn"):
        if not csv_file:
            st.warning("请先上传 CSV 文件。")
        else:
            text_ = csv_file.getvalue().decode("utf-8-sig", errors="ignore")
            reader = csv.reader(io.StringIO(text_))
            imported = 0
            with db_session() as db:
                for r in reader:
                    if not r or len(r) < 3:
                        continue
                    if str(r[0]).strip() in ("name", "城市", "city"):
                        continue
                    nm = str(r[0]).strip()
                    if not nm:
                        continue
                    try:
                        lat_v = float(r[1])
                        lon_v = float(r[2])
                    except Exception:
                        continue
                    rec = db.query(City).filter(City.name == nm).first()
                    if rec:
                        rec.lat = lat_v
                        rec.lon = lon_v
                    else:
                        db.add(City(name=nm, lat=lat_v, lon=lon_v))
                    imported += 1
                if not safe_commit(db, context="城市坐标 CSV 导入"):
                    st.stop()
            clear_runtime_caches_after_data_change()
            st.success(f"已导入 / 更新 {imported} 条城市坐标。")
            st.rerun()

    with db_session() as db:
        cities = db.query(City).order_by(City.id.desc()).limit(300).all()
    rows = [{"ID": c.id, "城市": c.name, "lat": round(float(c.lat), 6), "lon": round(float(c.lon), 6)} for c in cities]
    show_table(rows)
    if cities:
        delete_id = st.selectbox("删除城市（按ID）", [c.id for c in cities], key="delete_city_select")
        if st.button("删除所选城市", key="delete_city_btn"):
            with db_session() as db:
                obj = db.query(City).filter(City.id == delete_id).first()
                if obj:
                    db.delete(obj)
                    if not safe_commit(db, context=f"删除城市：city#{delete_id}"):
                        st.stop()
            clear_runtime_caches_after_data_change()
            st.success("已删除")
            st.rerun()

# -------------------- 模板导入 --------------------
elif page == "模板导入":
    st.subheader("模板导入")
    st.caption("下载模板 → 填写 → 上传导入，支持新增/更新。")

    def make_xlsx_template(headers, example_rows, sheet_name="template"):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(headers)
        for row in example_rows:
            ws.append(row)
        for i, h in enumerate(headers, start=1):
            col_letter = chr(64 + i) if i <= 26 else None
            if col_letter:
                ws.column_dimensions[col_letter].width = max(12, min(36, len(str(h)) * 2))
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio

    def read_xlsx_rows(uploaded_file):
        from openpyxl import load_workbook

        data = uploaded_file.getvalue()
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], []
        headers = [str(x).strip() if x is not None else "" for x in rows[0]]
        out = []
        for r in rows[1:]:
            if not r or all(x is None or str(x).strip() == "" for x in r):
                continue
            first = str(r[0]).strip() if r[0] is not None else ""
            if first in ("必填", "说明", "字段说明"):
                continue
            out.append(list(r))
        return headers, out

    def find_idx(headers, aliases: list[str]) -> Optional[int]:
        for cand in aliases:
            for i, h in enumerate(headers):
                if str(h).strip() == str(cand).strip():
                    return i
        for cand in aliases:
            for i, h in enumerate(headers):
                if str(h).strip().startswith(str(cand).strip()):
                    return i
        return None

    headers_a = [
        "姓名", "性别(男/女)", "等级(A/B/C)", "可带队(是/否)", "常驻城市", "每周上限(院次)",
        "状态(在岗/请假/冻结)", "本月已排院次", "本月差旅天数", "连续工作天数",
        "上次结束城市(可空)", "上次结束日期(YYYY-MM-DD)(必填)",
    ]
    explain_a = ["必填", "默认女", "默认B", "默认是", "必填", "默认1", "默认在岗", "默认0", "默认0", "默认0", "可空", "必填"]
    example_a = [
        ["张三", "女", "A", "是", "北京", 1, "在岗", 0, 0, 0, "苏州", "2026-01-20"],
        ["李四", "女", "B", "是", "上海", 2, "在岗", 0, 0, 0, "", "2026-02-01"],
    ]
    bio_a = make_xlsx_template(headers_a, [explain_a] + example_a, sheet_name="稽查员")
    st.download_button("下载稽查员模板（XLSX）", bio_a.getvalue(), file_name="稽查员模板.xlsx", key="dl_aud_tpl")

    headers_t = [
        "项目名称", "客户/申办方", "需要A带队(是/否)", "所需人数", "任务天数", "性别要求(男/女/不限)",
        "硬指定人员(可空)", "软指定专家/老师(可空)", "中心城市", "开始日期(YYYY-MM-DD)", "结束日期(YYYY-MM-DD)(必填)",
    ]
    explain_t = ["必填", "可空", "默认否", "默认1", "默认1", "默认不限", "可空", "可空", "必填", "必填", "必填"]
    example_t = [
        ["项目A", "申办方X", "否", 2, 2, "不限", "", "张三", "苏州", "2026-02-01", "2026-02-02"],
        ["项目B", "申办方Y", "是", 1, 3, "女", "", "", "北京", "2026-02-03", "2026-02-05"],
    ]
    bio_t = make_xlsx_template(headers_t, [explain_t] + example_t, sheet_name="任务")
    st.download_button("下载任务模板（XLSX）", bio_t.getvalue(), file_name="任务模板.xlsx", key="dl_task_tpl")

    st.divider()
    auditor_xlsx = st.file_uploader("上传稽查员模板", type=["xlsx"], key="auditor_xlsx")
    if st.button("导入稽查员模板", key="import_aud_btn"):
        if not auditor_xlsx:
            st.warning("请先上传稽查员模板。")
        else:
            headers, rows = read_xlsx_rows(auditor_xlsx)
            imported = 0
            with db_session() as db:
                for r in rows:
                    name_i = find_idx(headers, ["姓名"])
                    base_i = find_idx(headers, ["常驻城市"])
                    if name_i is None or base_i is None:
                        continue
                    name = str(r[name_i] or "").strip()
                    base_city = str(r[base_i] or "").strip()
                    if not name or not base_city:
                        continue

                    gender_idx = find_idx(headers, ["性别(男/女)"])
                    group_idx = find_idx(headers, ["等级(A/B/C)"])
                    lead_idx = find_idx(headers, ["可带队(是/否)"])
                    week_idx = find_idx(headers, ["每周上限(院次)"])
                    status_idx = find_idx(headers, ["状态(在岗/请假/冻结)"])
                    month_idx = find_idx(headers, ["本月已排院次"])
                    travel_idx = find_idx(headers, ["本月差旅天数"])
                    cont_idx = find_idx(headers, ["连续工作天数"])
                    last_city_idx = find_idx(headers, ["上次结束城市(可空)"])
                    last_date_idx = find_idx(headers, ["上次结束日期(YYYY-MM-DD)(必填)"])

                    gender = str(r[gender_idx] if gender_idx is not None else "女")
                    group_level = str(r[group_idx] if group_idx is not None else "B")
                    can_lead_raw = str(r[lead_idx] if lead_idx is not None else "是")
                    last_date_raw = r[last_date_idx] if last_date_idx is not None else None
                    last_date = safe_parse_date(last_date_raw) or date.today()

                    rec = db.query(Auditor).filter(Auditor.name == name).first()
                    if rec:
                        rec.gender = gender or "女"
                        rec.group_level = group_level or "B"
                        rec.can_lead_team = can_lead_raw in BOOL_TRUE
                        rec.base_city = base_city
                        rec.max_weekly_tasks = _safe_int(r[week_idx], 1) if week_idx is not None else 1
                        rec.status = STATUS_MAP.get(str(r[status_idx]).strip(), "active") if status_idx is not None else "active"
                        rec.monthly_cases = _safe_int(r[month_idx], 0) if month_idx is not None else 0
                        rec.travel_days = _safe_int(r[travel_idx], 0) if travel_idx is not None else 0
                        rec.continuous_days = _safe_int(r[cont_idx], 0) if cont_idx is not None else 0
                        rec.last_task_end_city = str(r[last_city_idx]).strip() if last_city_idx is not None and r[last_city_idx] is not None else None
                        rec.last_task_end_date = last_date
                    else:
                        db.add(
                            Auditor(
                                name=name,
                                gender=gender or "女",
                                group_level=group_level or "B",
                                can_lead_team=can_lead_raw in BOOL_TRUE,
                                base_city=base_city,
                                max_weekly_tasks=_safe_int(r[week_idx], 1) if week_idx is not None else 1,
                                status=STATUS_MAP.get(str(r[status_idx]).strip(), "active") if status_idx is not None else "active",
                                monthly_cases=_safe_int(r[month_idx], 0) if month_idx is not None else 0,
                                travel_days=_safe_int(r[travel_idx], 0) if travel_idx is not None else 0,
                                continuous_days=_safe_int(r[cont_idx], 0) if cont_idx is not None else 0,
                                last_task_end_city=str(r[last_city_idx]).strip() if last_city_idx is not None and r[last_city_idx] is not None else None,
                                last_task_end_date=last_date,
                            )
                        )
                    imported += 1
                if not safe_commit(db, "导入稽查员模板"):
                    st.stop()
            clear_runtime_caches_after_data_change()
            st.success(f"已导入 / 更新 {imported} 条稽查员记录。")
            st.rerun()

    st.divider()
    task_xlsx = st.file_uploader("上传任务模板", type=["xlsx"], key="task_xlsx")
    if st.button("导入任务模板", key="import_task_btn"):
        if not task_xlsx:
            st.warning("请先上传任务模板。")
        else:
            headers, rows = read_xlsx_rows(task_xlsx)
            imported = 0
            with db_session() as db:
                for r in rows:
                    proj_i = find_idx(headers, ["项目名称"])
                    city_i = find_idx(headers, ["中心城市"])
                    sd_i = find_idx(headers, ["开始日期(YYYY-MM-DD)"])
                    ed_i = find_idx(headers, ["结束日期(YYYY-MM-DD)(必填)"])
                    if None in (proj_i, city_i, sd_i, ed_i):
                        continue

                    project_name = str(r[proj_i] or "").strip()
                    site_city = str(r[city_i] or "").strip()
                    start_d = safe_parse_date(r[sd_i])
                    end_d = safe_parse_date(r[ed_i])

                    if not project_name or not site_city or not start_d or not end_d:
                        continue
                    if end_d < start_d:
                        continue

                    customer_i = find_idx(headers, ["客户/申办方"])
                    need_i = find_idx(headers, ["需要A带队(是/否)"])
                    head_i = find_idx(headers, ["所需人数"])
                    days_i = find_idx(headers, ["任务天数"])
                    gender_i = find_idx(headers, ["性别要求(男/女/不限)"])
                    hard_i = find_idx(headers, ["硬指定人员(可空)"])
                    soft_i = find_idx(headers, ["软指定专家/老师(可空)"])

                    rec = db.query(Task).filter(Task.project_name == project_name, Task.start_date == start_d, Task.site_city == site_city).first()
                    if rec:
                        rec.customer_name = str(r[customer_i]).strip() if customer_i is not None and r[customer_i] is not None else None
                        rec.need_expert = str(r[need_i]).strip() in BOOL_TRUE if need_i is not None else False
                        rec.required_headcount = _safe_int(r[head_i], 1) if head_i is not None else 1
                        rec.required_days = _safe_int(r[days_i], max(1, (end_d - start_d).days + 1)) if days_i is not None else max(1, (end_d - start_d).days + 1)
                        rec.required_gender = str(r[gender_i]).strip() if gender_i is not None and r[gender_i] is not None else "不限"
                        rec.specified_auditors = str(r[hard_i]).strip() if hard_i is not None and r[hard_i] is not None else None
                        rec.preferred_experts = str(r[soft_i]).strip() if soft_i is not None and r[soft_i] is not None else None
                        rec.end_date = end_d
                    else:
                        db.add(
                            Task(
                                project_name=project_name,
                                customer_name=str(r[customer_i]).strip() if customer_i is not None and r[customer_i] is not None else None,
                                need_expert=str(r[need_i]).strip() in BOOL_TRUE if need_i is not None else False,
                                required_headcount=_safe_int(r[head_i], 1) if head_i is not None else 1,
                                required_days=_safe_int(r[days_i], max(1, (end_d - start_d).days + 1)) if days_i is not None else max(1, (end_d - start_d).days + 1),
                                required_gender=str(r[gender_i]).strip() if gender_i is not None and r[gender_i] is not None else "不限",
                                specified_auditors=str(r[hard_i]).strip() if hard_i is not None and r[hard_i] is not None else None,
                                preferred_experts=str(r[soft_i]).strip() if soft_i is not None and r[soft_i] is not None else None,
                                site_city=site_city,
                                start_date=start_d,
                                end_date=end_d,
                            )
                        )
                    imported += 1
                if not safe_commit(db, "导入任务模板"):
                    st.stop()
            clear_runtime_caches_after_data_change()
            st.success(f"已导入 / 更新 {imported} 条任务记录。")
            st.rerun()



def aggregate_month_detail_rows_for_display(rows: list[dict]) -> list[dict]:
    grouped = {}
    for r in rows or []:
        key = (
            int(r.get("任务ID") or 0),
            r.get("项目", "") or "",
            r.get("城市", "") or "",
            r.get("开始日期", "") or "",
            r.get("结束日期", "") or "",
        )
        grouped.setdefault(key, {"sources": [], "roles": [], "people": [], "record_ids": []})
        source_label = r.get("来源", "") or ""
        role = r.get("角色", "") or ""
        person = r.get("人员", "") or ""
        record_id = r.get("记录ID")

        if source_label and source_label not in grouped[key]["sources"]:
            grouped[key]["sources"].append(source_label)
        if role:
            for one_role in [x.strip() for x in str(role).split("、") if x.strip()]:
                if one_role not in grouped[key]["roles"]:
                    grouped[key]["roles"].append(one_role)
        if person:
            for one_person in [x.strip() for x in str(person).split("、") if x.strip()]:
                if one_person not in grouped[key]["people"]:
                    grouped[key]["people"].append(one_person)
        if record_id is not None:
            grouped[key]["record_ids"].append(record_id)

    out = []
    for (task_id, project_name, city, start_s, end_s), payload in grouped.items():
        out.append({
            "来源": "、".join(payload["sources"]),
            "记录ID": "、".join(str(x) for x in payload["record_ids"]),
            "任务ID": task_id,
            "项目": project_name,
            "城市": city,
            "角色": "、".join(payload["roles"]),
            "人员": "、".join(payload["people"]),
            "开始日期": start_s,
            "结束日期": end_s,
        })
    out.sort(key=lambda x: (x["开始日期"], x["任务ID"], x["项目"]))
    return out


def load_month_schedule_detail_rows(year: int, month: int):
    month_start = date(int(year), int(month), 1)
    next_month = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
    month_end = next_month - timedelta(days=1)

    rows = []
    existing_schedule_keys = set()

    with db_session() as db:
        schedules = (
            db.query(Schedule)
            .options(joinedload(Schedule.task), joinedload(Schedule.auditor))
            .filter(Schedule.start_date <= month_end, Schedule.end_date >= month_start)
            .order_by(Schedule.start_date.asc(), Schedule.id.asc())
            .all()
        )
        for s in schedules:
            if not s.task:
                continue
            person_name = s.auditor.name if s.auditor else ""
            key = (int(s.task_id), person_name, d2s(s.start_date), d2s(s.end_date), s.role)
            existing_schedule_keys.add(key)
            rows.append({
                "来源": "标准排班",
                "记录ID": int(s.id),
                "任务ID": int(s.task_id),
                "项目": s.task.project_name if s.task else "",
                "城市": s.task.site_city if s.task else "",
                "角色": "组长" if s.role == "leader" else "成员",
                "人员": person_name,
                "开始日期": d2s(s.start_date),
                "结束日期": d2s(s.end_date),
                "_source_type": "schedule",
            })

    try:
        with engine.begin() as conn:
            drows = conn.execute(text("""
                SELECT id, task_id, auditor_id, person_name, role, start_date, end_date
                FROM direct_assignments
                WHERE start_date <= :month_end AND end_date >= :month_start
                ORDER BY start_date ASC, id ASC
            """), {"month_start": d2s(month_start), "month_end": d2s(month_end)}).mappings().all()

        with db_session() as db:
            task_map = {int(t.id): t for t in db.query(Task).all()}

        for r in drows:
            if r.get("task_id") is None:
                continue
            task_obj = task_map.get(int(r["task_id"]))
            if not task_obj:
                continue
            person_name = r.get("person_name") or ""
            key = (int(r["task_id"]), person_name, str(r.get("start_date") or ""), str(r.get("end_date") or ""), r.get("role") or "")
            if key in existing_schedule_keys:
                continue
            rows.append({
                "来源": "已定直录",
                "记录ID": int(r["id"]),
                "任务ID": int(r["task_id"]),
                "项目": task_obj.project_name,
                "城市": task_obj.site_city,
                "角色": "组长" if r.get("role") == "leader" else "成员",
                "人员": person_name,
                "开始日期": str(r.get("start_date") or ""),
                "结束日期": str(r.get("end_date") or ""),
                "_source_type": "direct",
            })
    except Exception:
        pass

    rows.sort(key=lambda x: (x["开始日期"], x["任务ID"], x["来源"], x["记录ID"]))
    return rows

def delete_schedule_detail_row(source_type: str, record_id: int, task_id: int):
    if source_type == "schedule":
        with db_session() as db:
            db.query(Schedule).filter(Schedule.id == int(record_id)).delete()
            return safe_commit(db, f"删除排班记录#{record_id}")
    if source_type == "direct":
        ensure_support_tables()
        with engine.begin() as conn:
            conn.execute(text("DELETE FROM direct_assignments WHERE id = :id"), {"id": int(record_id)})
        with db_session() as db:
            task = db.query(Task).filter(Task.id == int(task_id)).first()
            if task:
                db.query(Schedule).filter(Schedule.task_id == int(task_id)).delete()
                safe_commit(db, f"同步清理任务#{task_id}标准排班")
                sync_task_schedules_from_direct_assignments(task)
        return True
    return False

def update_schedule_detail_row(source_type: str, record_id: int, task_id: int, role_label: str, person_name: str, start_date_value, end_date_value):
    start_d = safe_parse_date(start_date_value)
    end_d = safe_parse_date(end_date_value)
    if not start_d or not end_d:
        return False, "开始/结束日期无效"
    if end_d < start_d:
        return False, "结束日期不能早于开始日期"

    role_value = "leader" if str(role_label) == "组长" else "member"

    if source_type == "schedule":
        with db_session() as db:
            row = db.query(Schedule).options(joinedload(Schedule.auditor)).filter(Schedule.id == int(record_id)).first()
            if not row:
                return False, "未找到排班记录"
            auditor = row.auditor
            if person_name and auditor and auditor.name != person_name:
                new_auditor = db.query(Auditor).filter(Auditor.name == person_name).first()
                if new_auditor:
                    row.auditor_id = int(new_auditor.id)
            row.role = role_value
            row.start_date = start_d
            row.end_date = end_d
            ok = safe_commit(db, f"更新排班记录#{record_id}")
            return (ok, "已更新") if ok else (False, "更新失败")

    if source_type == "direct":
        ensure_support_tables()
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with engine.begin() as conn:
            conn.execute(text("""
                UPDATE direct_assignments
                SET person_name=:person_name, role=:role, start_date=:start_date, end_date=:end_date, updated_at=:updated_at
                WHERE id=:id
            """), {
                "person_name": person_name,
                "role": role_value,
                "start_date": d2s(start_d),
                "end_date": d2s(end_d),
                "updated_at": now,
                "id": int(record_id),
            })
        with db_session() as db:
            task = db.query(Task).filter(Task.id == int(task_id)).first()
            if task:
                db.query(Schedule).filter(Schedule.task_id == int(task_id)).delete()
                safe_commit(db, f"清理任务#{task_id}旧排班")
                sync_task_schedules_from_direct_assignments(task)
        return True, "已更新"
    return False, "未知来源"

# -------------------- 日历视图 --------------------
if page == "日历视图":
    st.subheader("日历视图")
    st.caption("按月查看排班、节假日标识，并支持导出 ICS 日历。")

    c1, c2, c3 = st.columns(3)
    with db_session() as db:
        auditors = db.query(Auditor).order_by(Auditor.name.asc()).all()

    auditor_options = {"全部稽查员": None}
    for a in auditors:
        auditor_options[f"#{a.id} {a.name}"] = a.id

    auditor_label = c1.selectbox("筛选稽查员", list(auditor_options.keys()), key="cal_auditor_filter")
    year = c2.selectbox("年份", list(range(date.today().year - 2, date.today().year + 3)), index=2, key="cal_year")
    month = c3.selectbox("月份", list(range(1, 13)), index=date.today().month - 1, key="cal_month")
    auditor_id = auditor_options[auditor_label]

    month_start = date(int(year), int(month), 1)
    next_month = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
    month_end = next_month - timedelta(days=1)

    with db_session() as db:
        all_schedules = (
            db.query(Schedule)
            .options(joinedload(Schedule.task), joinedload(Schedule.auditor))
            .filter(Schedule.start_date <= month_end, Schedule.end_date >= month_start)
            .order_by(Schedule.start_date.asc())
            .all()
        )

    direct_rows_raw = []
    ensure_extra_tables()
    with engine.begin() as conn:
        rows = conn.execute(
            text("""
                SELECT da.id, da.task_id, da.auditor_id, da.person_name, da.is_part_time, da.role, da.start_date, da.end_date, da.notes,
                       t.project_name, t.site_city
                FROM direct_assignments da
                LEFT JOIN tasks t ON da.task_id = t.id
                WHERE da.start_date <= :month_end AND da.end_date >= :month_start
                ORDER BY da.start_date ASC, da.person_name ASC
            """),
            {"month_start": d2s(month_start), "month_end": d2s(month_end)},
        ).mappings().all()
        direct_rows_raw = [dict(r) for r in rows]

    direct_task_ids = {int(r["task_id"]) for r in direct_rows_raw}
    all_schedules_rows = []

    for s in all_schedules:
        if int(s.task_id) in direct_task_ids:
            continue
        if auditor_id and s.auditor_id != auditor_id:
            continue
        all_schedules_rows.append(
            {
                "id": s.id,
                "auditor_id": s.auditor_id,
                "auditor_name": (s.auditor.name if s.auditor else ""),
                "task_id": s.task_id,
                "project_name": (s.task.project_name if s.task else ""),
                "site_city": (s.task.site_city if s.task else ""),
                "role": s.role,
                "start_date": s.start_date,
                "end_date": s.end_date,
                "travel_from_city": s.travel_from_city,
                "travel_to_city": s.travel_to_city,
                "distance_km": float(s.distance_km or 0),
                "source": "schedule",
            }
        )

    direct_rows = []
    for r in direct_rows_raw:
        if auditor_id and r.get("auditor_id") and int(r.get("auditor_id")) != int(auditor_id):
            continue
        direct_rows.append(
            {
                "id": r.get("id"),
                "auditor_id": r.get("auditor_id"),
                "auditor_name": r.get("person_name") or "",
                "task_id": r.get("task_id"),
                "project_name": r.get("project_name") or "",
                "site_city": r.get("site_city") or "",
                "role": r.get("role") or "member",
                "start_date": safe_parse_date(r.get("start_date")),
                "end_date": safe_parse_date(r.get("end_date")),
                "travel_from_city": "",
                "travel_to_city": r.get("site_city") or "",
                "distance_km": 0.0,
                "source": "direct",
            }
        )

    merged_rows = all_schedules_rows + direct_rows
    calendar_detail_map = build_calendar_detail_map(merged_rows)
    day_marks = {it.get("date"): it for it in load_day_marks() if it.get("date", "")[:7] == month_start.strftime("%Y-%m")}

    events_by_day = {}
    for s in merged_rows:
        cur = s.get("start_date")
        end_d = s.get("end_date")
        while cur and end_d and cur <= end_d:
            key = (cur.isoformat(), int(s.get("task_id")))
            events_by_day.setdefault(key, {"project": s.get("project_name") or f"任务#{s.get('task_id')}", "task_id": s.get("task_id"), "persons": [], "city": s.get("site_city") or ""})
            nm = s.get("auditor_name") or ""
            if nm and nm not in events_by_day[key]["persons"]:
                events_by_day[key]["persons"].append(nm)
            cur += timedelta(days=1)

    st.subheader(f"{year}年{month}月 日历总览")
    weeks = []
    first_cell = month_start - timedelta(days=month_start.weekday())
    current = first_cell
    for _ in range(6):
        row = []
        for _ in range(7):
            row.append(current)
            current += timedelta(days=1)
        weeks.append(row)

    headers = st.columns(7)
    for idx, h in enumerate(["周一", "周二", "周三", "周四", "周五", "周六", "周日"]):
        headers[idx].markdown(f"**{h}**")

    for week in weeks:
        cols = st.columns(7)
        for idx, day in enumerate(week):
            marks = []
            mk = day_marks.get(day.isoformat())
            if mk:
                marks.append(mk.get("label") or mk.get("type") or "标记")

            evs = []
            for (day_iso, task_id), obj in events_by_day.items():
                if day_iso != day.isoformat():
                    continue
                evs.append(f"#{task_id} {obj['project']}｜{'、'.join(obj['persons'])}")

            color = "#ffffff"
            if day.month != month:
                color = "#f7f7f7"
            elif evs:
                color = "#eef6ff"

            with cols[idx]:
                with st.container(border=True):
                    trigger_label = f"{day.day}"
                    if marks:
                        trigger_label += " " + " / ".join(marks)
                    with st.popover(trigger_label, use_container_width=True):
                        render_calendar_day_popover(day, calendar_detail_map.get(day, []))
                    if evs:
                        for e in evs[:3]:
                            st.caption(e)
                        if len(evs) > 3:
                            st.caption(f"还有 {len(evs)-3} 项，点击日期查看全部")

    render_day_detail_panel(calendar_detail_map)

    st.divider()
    st.subheader("本月排班明细")
    month_detail_rows = load_month_schedule_detail_rows(int(year), int(month))
    if month_detail_rows:
        month_detail_display_rows = aggregate_month_detail_rows_for_display(month_detail_rows)
        st.dataframe(
            pd.DataFrame(month_detail_display_rows),
            use_container_width=True,
            height=420,
            hide_index=True,
        )

        row_options = {
            f"{r['来源']}｜记录{r['记录ID']}｜任务{r['任务ID']}｜{r['项目']}｜{r['人员']}": (r["_source_type"], r["记录ID"], r["任务ID"])
            for r in month_detail_rows
        }
        selected_label = st.selectbox("选择要修改或删除的排班明细", list(row_options.keys()), key="calendar_detail_edit_select")
        selected_source, selected_record_id, selected_task_id = row_options[selected_label]
        selected_row = next(r for r in month_detail_rows if r["记录ID"] == selected_record_id and r["_source_type"] == selected_source)

        with st.expander("修改 / 删除本条排班明细", expanded=True):
            c1, c2, c3, c4 = st.columns(4)
            edit_person = c1.text_input("人员", value=selected_row["人员"], key=f"edit_person_{selected_source}_{selected_record_id}")
            edit_role = c2.selectbox("角色", ["组长", "成员"], index=0 if selected_row["角色"] == "组长" else 1, key=f"edit_role_{selected_source}_{selected_record_id}")
            edit_start = c3.text_input("开始日期", value=selected_row["开始日期"], key=f"edit_start_{selected_source}_{selected_record_id}")
            edit_end = c4.text_input("结束日期", value=selected_row["结束日期"], key=f"edit_end_{selected_source}_{selected_record_id}")

            b1, b2 = st.columns(2)
            if b1.button("保存本条修改", key=f"save_sched_row_{selected_source}_{selected_record_id}", type="primary"):
                ok, msg = update_schedule_detail_row(
                    selected_source, selected_record_id, selected_task_id, edit_role, edit_person, edit_start, edit_end
                )
                if ok:
                    clear_runtime_caches_after_data_change()
                    st.success(msg)
                    st.rerun()
                else:
                    st.error(msg)

            if b2.button("删除本条排班明细", key=f"delete_sched_row_{selected_source}_{selected_record_id}"):
                ok = delete_schedule_detail_row(selected_source, selected_record_id, selected_task_id)
                if ok:
                    clear_runtime_caches_after_data_change()
                    st.success("已删除")
                    st.rerun()
                else:
                    st.error("删除失败")
    else:
        st.info("本月暂无排班明细")


# -------------------- 账号管理 --------------------
# -------------------- 账号管理 --------------------
elif page == "账号管理":
    st.subheader("账号管理")
    current_user = st.session_state.get("login_user", "")
    is_admin = bool(st.session_state.get("is_admin", False))
    is_super_admin = bool(st.session_state.get("is_super_admin", False))

    st.subheader("我的密码")
    with st.form("change_my_password", clear_on_submit=True):
        old_pw = st.text_input("当前密码", type="password")
        new_pw = st.text_input("新密码（至少6位）", type="password")
        new_pw2 = st.text_input("确认新密码", type="password")
        if st.form_submit_button("修改我的密码", type="primary"):
            if not check_login(current_user, old_pw):
                st.error("当前密码不正确")
            elif new_pw != new_pw2:
                st.error("两次输入的新密码不一致")
            else:
                ok, msg = update_auth_password(current_user, new_pw)
                if ok:
                    st.success(msg)
                else:
                    st.error(msg)

    st.divider()
    if not is_admin:
        st.info("当前账号仅可修改自己的密码。新增登录人员、重置他人密码、配置可见板块仅管理员可操作。")
    else:
        st.subheader("新增登录人员")
        with st.form("create_user_form", clear_on_submit=True):
            c1, c2, c3 = st.columns(3)
            new_username = c1.text_input("新账号")
            new_password = c2.text_input("初始密码（至少6位）", type="password")
            role = c3.selectbox("权限", ["普通用户", "管理员", "主管理员"])
            if st.form_submit_button("新增账号", type="primary"):
                ok, msg = create_auth_user(
                    new_username,
                    new_password,
                    is_admin=(role in ("管理员", "主管理员")),
                    is_super_admin=(role == "主管理员"),
                )
                if ok:
                    st.success(msg)
                    st.rerun()
                else:
                    st.error(msg)

        st.subheader("现有登录账号")
        users = list_auth_users()
        if users:
            rows = []
            for u in users:
                role_cn = "普通用户"
                if int(u.get("is_super_admin", 0)) == 1:
                    role_cn = "主管理员"
                elif int(u.get("is_admin", 0)) == 1:
                    role_cn = "管理员"
                rows.append({"账号": u.get("username"), "权限": role_cn, "创建时间": u.get("created_at") or ""})
            show_table(rows, 260)
        else:
            st.info("暂无账号")

        st.subheader("重置其他人员密码")
        user_labels = [u["username"] for u in users]
        if user_labels:
            with st.form("reset_password_form", clear_on_submit=True):
                c1, c2 = st.columns(2)
                reset_user = c1.selectbox("选择账号", user_labels)
                reset_pw = c2.text_input("新密码（至少6位）", type="password")
                if st.form_submit_button("重置密码"):
                    ok, msg = update_auth_password(reset_user, reset_pw)
                    if ok:
                        st.success(f"{reset_user}：{msg}")
                    else:
                        st.error(msg)

            st.subheader("删除登录账号")
            deletable = [u for u in user_labels if u not in ("admin", current_user)]
            if deletable:
                with st.form("delete_user_form", clear_on_submit=True):
                    del_user = st.selectbox("选择要删除的账号", deletable)
                    confirm_text = st.text_input("输入 DELETE 确认删除")
                    if st.form_submit_button("删除账号"):
                        if confirm_text != "DELETE":
                            st.error("请输入 DELETE 以确认删除")
                        else:
                            ok, msg = delete_auth_user(del_user, current_user)
                            if ok:
                                st.success(msg)
                                st.rerun()
                            else:
                                st.error(msg)

        st.divider()
        if not is_super_admin:
            st.info("提示：只有【主管理员】可以配置普通账号的可见板块。")
        else:
            st.subheader("普通账号可见板块配置（主管理员）")
            st.caption("勾选后保存：普通账号侧边栏仅显示被勾选的功能。管理员/主管理员默认全功能，不受此限制。")

            normal_users = []
            for u in users:
                if int(u.get("is_admin", 0)) == 1:
                    continue
                normal_users.append(u.get("username"))

            if not normal_users:
                st.info("暂无普通账号")
            else:
                target_user = st.selectbox("选择普通账号", normal_users, key="perm_target_user")
                current_pages = get_user_allowed_pages(target_user)
                selected_pages = st.multiselect(
                    "可见板块（勾选）",
                    options=ALL_PAGES,
                    default=current_pages,
                    key="perm_pages_multiselect",
                )
                c1, _ = st.columns([1, 3])
                if c1.button("保存可见板块", type="primary", key="save_perm_btn"):
                    ok, msg = set_user_allowed_pages(target_user, selected_pages)
                    if ok:
                        st.success(msg)
                        if str(target_user).strip() == str(current_user).strip():
                            st.session_state["allowed_pages"] = get_user_allowed_pages(current_user)
                        st.rerun()
                    else:
                        st.error(msg)

# -------------------- 数据清理 --------------------
elif page == "数据清理":
    st.subheader("数据清理")
    st.warning("当前无数据时，可直接清空所有业务表。此操作不可恢复。")
    with st.form("cleanup_form"):
        confirm = st.text_input("输入 CLEAR 确认清空")
        submitted = st.form_submit_button("清空全部业务数据", type="primary")
    if submitted:
        if confirm != "CLEAR":
            st.error("请输入 CLEAR")
        else:
            with db_session() as db:
                db.query(Schedule).delete()
                db.query(Task).delete()
                db.query(Auditor).delete()
                db.query(CityDistance).delete()
                db.query(City).delete()
                if safe_commit(db, "清空业务数据"):
                    clear_runtime_caches_after_data_change()
                    st.success("已清空")
                    st.rerun()

else:
    st.info("请选择左侧功能导航。")
