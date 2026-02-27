import io
import re
import json
from pathlib import Path
from fastapi import FastAPI, Depends, Request, Form, UploadFile, HTTPException, File, Query
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import func
from .db import Base, engine, get_db
from .seed_distances import SEED_CITY_DISTANCES
from .models import Auditor, Task, Schedule, CityDistance
from .scheduler import build_candidates, propose_team, compute_from_city, get_distance_km, iso_week, team_objective
from datetime import datetime, date, timedelta

from urllib.parse import quote
from typing import Optional

def _content_disposition_utf8(filename_utf8: str, ascii_fallback: str = "template.xlsx") -> str:
    """Starlette 响应头要求 latin-1，可用 filename* 兼容中文文件名。"""
    q = quote(filename_utf8)
    return f'attachment; filename="{ascii_fallback}"; filename*=UTF-8\'\'{q}'


Base.metadata.create_all(bind=engine)
from .db import ensure_schema
ensure_schema()

app = FastAPI(title="Audit Scheduler V1.3")
templates = Jinja2Templates(directory="app/templates")

# ---- Import Templates (XLSX) ----
def make_xlsx_template(headers, example_rows, sheet_name="template"):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in example_rows:
        ws.append(row)
    # set column widths
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[chr(64+i)].width = max(12, min(28, len(str(h))*2))
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

def read_xlsx_rows(upload_file):
    from openpyxl import load_workbook
    data = upload_file.file.read()
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(x).strip() if x is not None else "" for x in rows[0]]
    out = []
    for r in rows[1:]:
        if not r or all(x is None or str(x).strip()=="" for x in r):
            continue
        first = str(r[0]).strip() if r[0] is not None else ""
        if first in ("必填","说明","字段说明"):
            continue
        out.append(list(r))
    return headers, out



def seed_city_distances_if_empty(db: Session):
    """若距离表为空，则写入内置城市直线距离（一次性初始化）。
    注意：内置表可能存在重复城市对，需去重并忽略冲突，避免启动时报 500。
    """
     # 允许数据库已有部分数据：仍补齐缺失的内置距离（不会覆盖人工维护的 km）
    
    seen = set()
    for a, b, km in SEED_CITY_DISTANCES:
        key = (str(a).strip(), str(b).strip())
        if not key[0] or not key[1] or key[0] == key[1]:
            continue
        if key in seen:
            continue
        seen.add(key)
        # 若数据库已有（极少数并发/残留情况），则跳过
        exists = db.query(CityDistance).filter(CityDistance.from_city == key[0], CityDistance.to_city == key[1]).first()
        if exists:
            continue
        db.add(CityDistance(from_city=key[0], to_city=key[1], km=float(km)))
    try:
        db.commit()
    except Exception:
        db.rollback()
        # 兜底：逐条写入，忽略重复/冲突
        for a, b, km in SEED_CITY_DISTANCES:
            key = (str(a).strip(), str(b).strip())
            if not key[0] or not key[1] or key[0] == key[1]:
                continue
            exists = db.query(CityDistance).filter(CityDistance.from_city == key[0], CityDistance.to_city == key[1]).first()
            if exists:
                continue
            try:
                db.add(CityDistance(from_city=key[0], to_city=key[1], km=float(km)))
                db.commit()
            except Exception:
                db.rollback()
                continue

def parse_date(d: str):
    return datetime.strptime(d, "%Y-%m-%d").date()

@app.get("/", response_class=HTMLResponse)
def home(db: Session = Depends(get_db)):
    seed_city_distances_if_empty(db)
    return RedirectResponse(url="/schedule", status_code=302)

# ---- Auditors ----
@app.get("/auditors", response_class=HTMLResponse)
def auditors_page(request: Request, db: Session = Depends(get_db)):
    auditors = db.query(Auditor).order_by(Auditor.id.desc()).all()
    return templates.TemplateResponse("auditors.html", {"request": request, "auditors": auditors})

@app.post("/auditors/create")
def auditors_create(
    name: str = Form(...),
    gender: str = Form("男"),
    group_level: str = Form("B"),
    can_lead_team: str = Form("no"),
    base_city: str = Form(...),
    max_weekly_tasks: int = Form(1),
    monthly_cases: int = Form(0),
    travel_days: int = Form(0),
    continuous_days: int = Form(0),
    last_task_end_city: str = Form(""),
    last_task_end_date: str = Form(""),
    status: str = Form("active"),
    db: Session = Depends(get_db),
):
    a = Auditor(
        name=name.strip(),
        gender=gender,
        group_level=group_level,
        can_lead_team=(can_lead_team == "yes"),
        base_city=base_city.strip(),
        max_weekly_tasks=max_weekly_tasks,
        monthly_cases=monthly_cases,
        travel_days=travel_days,
        continuous_days=continuous_days,
        last_task_end_city=(last_task_end_city.strip() or None),
        last_task_end_date=(parse_date(last_task_end_date) if last_task_end_date.strip() else None),
        status=status,
    )
    db.add(a)
    db.commit()
    return RedirectResponse(url="/auditors", status_code=303)

@app.post("/auditors/delete/{auditor_id}")
def auditors_delete(auditor_id: int, db: Session = Depends(get_db)):
    a = db.query(Auditor).filter(Auditor.id == auditor_id).first()
    if a:
        db.delete(a); db.commit()
    return RedirectResponse(url="/auditors", status_code=303)

# ---- Tasks ----
@app.get("/tasks", response_class=HTMLResponse)
def tasks_page(request: Request, db: Session = Depends(get_db)):
    tasks = db.query(Task).order_by(Task.id.desc()).all()
    return templates.TemplateResponse("tasks.html", {"request": request, "tasks": tasks})

@app.post("/tasks/create")
def tasks_create(
    project_name: str = Form(...),
    customer_name: str = Form(""),
    need_expert: str = Form("no"),
    required_headcount: int = Form(1),
    required_days: int = Form(1),
    specified_auditors: str = Form(""),
    preferred_experts: str = Form(""),
    required_gender: str = Form("不限"),
    site_city: str = Form(...),
    start_date: str = Form(...),
    end_date: str = Form(""),
    db: Session = Depends(get_db),
):
    t = Task(
        project_name=project_name.strip(),
        customer_name=customer_name.strip(),
        need_expert=(need_expert == "yes"),
        required_headcount=max(1, int(required_headcount)),
        required_days=max(1, int(required_days)),
        specified_auditors=specified_auditors.strip() or None,
        preferred_experts=preferred_experts.strip() or None,
        required_gender=required_gender,
        site_city=site_city.strip(),
        start_date=parse_date(start_date),
        end_date=(parse_date(end_date) if end_date and end_date.strip() else None),
    )
    db.add(t); db.commit()
    return RedirectResponse(url="/tasks", status_code=303)

@app.post("/tasks/delete/{task_id}")
def tasks_delete(task_id: int, db: Session = Depends(get_db)):
    t = db.query(Task).filter(Task.id == task_id).first()
    if t:
        db.delete(t); db.commit()
    return RedirectResponse(url="/tasks", status_code=303)

# ---- Distances ----
@app.get("/distances", response_class=HTMLResponse)
def distances_page(request: Request, db: Session = Depends(get_db)):
    seed_city_distances_if_empty(db)
    dists = db.query(CityDistance).order_by(CityDistance.id.desc()).limit(200).all()
    return templates.TemplateResponse("distances.html", {"request": request, "dists": dists})

@app.post("/distances/upsert")
def distances_upsert(from_city: str = Form(...), to_city: str = Form(...), km: float = Form(...), db: Session = Depends(get_db)):
    fc, tc = from_city.strip(), to_city.strip()
    rec = db.query(CityDistance).filter(CityDistance.from_city == fc, CityDistance.to_city == tc).first()
    if rec:
        rec.km = float(km)
    else:
        db.add(CityDistance(from_city=fc, to_city=tc, km=float(km)))
    db.commit()
    return RedirectResponse(url="/distances", status_code=303)

@app.post("/distances/delete/{dist_id}")
def distances_delete(dist_id: int, db: Session = Depends(get_db)):
    rec = db.query(CityDistance).filter(CityDistance.id == dist_id).first()
    if rec:
        db.delete(rec); db.commit()
    return RedirectResponse(url="/distances", status_code=303)



# ---- Calendar Export (ICS) ----
def _ics_escape(s: str) -> str:
    return (s or "").replace("\\", "\\\\").replace(";", "\\;").replace(",", "\\,").replace("\n", "\\n")

def build_ics_events(db: Session, auditor_id: int = None):
    from .models import Schedule, Auditor, Task
    q = db.query(Schedule).order_by(Schedule.id.desc())
    if auditor_id:
        q = q.filter(Schedule.auditor_id == auditor_id)
    sch = q.all()
    events = []
    for s in sch:
        a = db.query(Auditor).filter(Auditor.id == s.auditor_id).first()
        t = db.query(Task).filter(Task.id == s.task_id).first()
        if not a or not t:
            continue
        start = datetime.datetime.combine(t.start_date, datetime.time(9,0,0))
        # ICS 的 DTEND 通常按“结束日期次日”表示（半开区间），这里用 18:00 作为结束时间。
        if getattr(t, 'end_date', None):
            _ed = t.end_date
            try:
                # 若 end_date < start_date，兜底回退
                if _ed < t.start_date:
                    _ed = t.start_date
            except Exception:
                _ed = t.start_date
            end_date = _ed + timedelta(days=1)
        else:
            end_date = t.start_date + timedelta(days=max(1, int(t.required_days or 1)))
        end = datetime.datetime.combine(end_date, datetime.time(18,0,0))
        uid = f"wnrh-{s.id}@scheduler"
        summary = f"{t.project_name}｜{t.site_city}｜{s.role}"
        desc = f"客户:{t.customer_name or ''}\n人数:{t.required_headcount} 天数:{t.required_days}\n负责人/成员:{a.name}"
        events.append((uid, start, end, summary, desc))
    return events

def make_ics(events, cal_name="万宁睿和排班"):
    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//WNRH Scheduler//CN",
        "CALSCALE:GREGORIAN",
        f"X-WR-CALNAME:{_ics_escape(cal_name)}",
    ]
    now = datetime.datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    for uid, start, end, summary, desc in events:
        lines += [
            "BEGIN:VEVENT",
            f"UID:{uid}",
            f"DTSTAMP:{now}",
            f"DTSTART:{start.strftime('%Y%m%dT%H%M%S')}",
            f"DTEND:{end.strftime('%Y%m%dT%H%M%S')}",
            f"SUMMARY:{_ics_escape(summary)}",
            f"DESCRIPTION:{_ics_escape(desc)}",
            "END:VEVENT",
        ]
    lines.append("END:VCALENDAR")
    return "\r\n".join(lines).encode("utf-8")

@app.get("/calendar", response_class=HTMLResponse)
def calendar_page(request: Request, db: Session = Depends(get_db)):
    from .models import Auditor
    auditors = db.query(Auditor).order_by(Auditor.name.asc()).all()
    return templates.TemplateResponse("calendar.html", {"request": request, "auditors": auditors})

@app.get("/api/calendar/events")
def api_calendar_events(
    start: Optional[str] = Query(None),
    end: Optional[str] = Query(None),
    auditor_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    """FullCalendar 事件源：返回 JSON 列表。
    - FullCalendar 通常会传 start/end（ISO 日期或日期时间）
    - 用户直接打开 /api/calendar/events 时可能不带 start/end：此时返回“当月范围”的事件，避免 422
    """
    # 兜底默认区间：当月 1 日 ~ 下月 1 日（含）
    today = date.today()
    month_start = date(today.year, today.month, 1)
    next_month_start = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
    default_start_d = month_start
    default_end_d = next_month_start

    def to_date(s: Optional[str], fallback: date) -> date:
        if not s:
            return fallback
        try:
            return date.fromisoformat(str(s)[:10])
        except Exception:
            return fallback

    start_d = to_date(start, default_start_d)
    end_d = to_date(end, default_end_d)

    q = (
        db.query(Schedule)
        .join(Task, Schedule.task_id == Task.id)
        .join(Auditor, Schedule.auditor_id == Auditor.id)
    )
    # overlap: schedule.start <= end and schedule.end >= start
    q = q.filter(Schedule.start_date <= end_d, Schedule.end_date >= start_d)
    if auditor_id:
        q = q.filter(Schedule.auditor_id == auditor_id)

    items = []
    for s in q.all():
        task = s.task
        auditor = s.auditor
        title = f"{task.project_name}｜{auditor.name}"
        # FullCalendar all-day end 为“次日”（end exclusive）
        end_exclusive = (s.end_date + timedelta(days=1)).isoformat()
        date_range = f"{s.start_date.isoformat()} ~ {s.end_date.isoformat()}"
        travel = f"{s.travel_from_city} → {s.travel_to_city}" if s.travel_from_city and s.travel_to_city else ""
        items.append({
            "id": s.id,
            "title": title,
            "start": s.start_date.isoformat(),
            "end": end_exclusive,
            "allDay": True,
            "extendedProps": {
                "project": getattr(task, "project_name", None) or getattr(task, "project", None) or "",
                "city": getattr(task, "site_city", "") or getattr(task, "city", "") or "",
                "role": s.role,
                "auditor": auditor.name if auditor else "",
                "date_range": date_range,
                "travel": travel,
                "distance_km": float(getattr(s, "distance_km", 0.0) or 0.0),
            }
        })

    return JSONResponse(items)

# ---------------- 日历：法定节假日/调休/补班 标识 ----------------
def _load_day_marks():
    """从 app/holidays_cn.json 读取日期标识。支持用户自行维护。"""
    try:
        p = Path(__file__).resolve().parent / "holidays_cn.json"
        if p.exists():
            obj = json.loads(p.read_text(encoding="utf-8"))
            items = obj.get("items", [])
            if isinstance(items, list):
                return items
    except Exception:
        pass
    return []

@app.get("/api/calendar/day-marks")
def api_calendar_day_marks(
    start: Optional[str] = Query(None),
    end: Optional[str] = Query(None),
):
    """返回区间内的日期标识：holiday(休)/workday(班)/adjust(调休)
    - 前端 FullCalendar 会传 start/end
    - 用户直接打开 /api/calendar/day-marks 时可能不带 start/end：此时返回“当月范围”，避免 422
    """
    today = date.today()
    month_start = date(today.year, today.month, 1)
    next_month_start = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
    default_start_d = month_start
    default_end_d = next_month_start

    def to_date(s: Optional[str], fallback: date) -> date:
        if not s:
            return fallback
        try:
            return date.fromisoformat(str(s)[:10])
        except Exception:
            return fallback

    start_d = to_date(start, default_start_d)
    end_d = to_date(end, default_end_d)

    marks = []
    for it in _load_day_marks():
        try:
            d = it.get("date", "")
            dt = date.fromisoformat(d)
            if start_d <= dt <= end_d:
                marks.append({
                    "date": d,
                    "type": it.get("type", ""),
                    "label": it.get("label", "")
                })
        except Exception:
            continue
    return JSONResponse(marks)

@app.get("/calendar/export/all.ics")
def export_calendar_all(db: Session = Depends(get_db)):
    data = make_ics(build_ics_events(db), cal_name="万宁睿和排班-全部")
    return StreamingResponse(
        io.BytesIO(data),
        media_type="text/calendar",
        headers={"Content-Disposition": 'attachment; filename="wnrh_all.ics"'},
    )


@app.get("/calendar/export/auditor/{auditor_id}.ics")
def export_calendar_auditor(auditor_id: int, db: Session = Depends(get_db)):
    from .models import Auditor
    a = db.query(Auditor).filter(Auditor.id == auditor_id).first()
    name = a.name if a else f"auditor_{auditor_id}"
    data = make_ics(build_ics_events(db, auditor_id=auditor_id), cal_name=f"万宁睿和排班-{name}")
    from urllib.parse import quote
    safe = f"auditor_{auditor_id}" if not a else f"auditor_{a.id}"
    fn_ascii = f"wnrh_{safe}.ics"
    fn_utf8 = quote(f"wnrh_{name}.ics")
    cd = f'attachment; filename="{fn_ascii}"; filename*=UTF-8\'\'{fn_utf8}'
    return StreamingResponse(io.BytesIO(data), media_type="text/calendar",
                             headers={"Content-Disposition": cd})


# ---- Cities (Coordinates) ----
SEED_CITIES = [
    ("北京", 39.9042, 116.4074), ("上海", 31.2304, 121.4737), ("广州", 23.1291, 113.2644), ("深圳", 22.5431, 114.0579),
    ("天津", 39.0851, 117.1994), ("重庆", 29.5630, 106.5516), ("成都", 30.5728, 104.0668), ("杭州", 30.2741, 120.1551),
    ("南京", 32.0603, 118.7969), ("苏州", 31.2989, 120.5853), ("武汉", 30.5928, 114.3055), ("西安", 34.3416, 108.9398),
    ("郑州", 34.7466, 113.6254), ("长沙", 28.2282, 112.9388), ("合肥", 31.8206, 117.2272), ("济南", 36.6512, 117.1201),
    ("青岛", 36.0671, 120.3826), ("大连", 38.9140, 121.6147), ("沈阳", 41.8057, 123.4315), ("厦门", 24.4798, 118.0894),
    ("福州", 26.0745, 119.2965), ("昆明", 25.0389, 102.7183), ("南宁", 22.8170, 108.3669), ("海口", 20.0444, 110.1983),
    ("石家庄", 38.0428, 114.5149), ("太原", 37.8706, 112.5489), ("南昌", 28.6829, 115.8582), ("贵阳", 26.6470, 106.6302),
    ("兰州", 36.0611, 103.8343), ("乌鲁木齐", 43.8256, 87.6168), ("呼和浩特", 40.8426, 111.7492), ("哈尔滨", 45.8038, 126.5349),
    ("长春", 43.8171, 125.3235), ("宁波", 29.8683, 121.5440), ("无锡", 31.4912, 120.3119), ("常州", 31.8107, 119.9741),

    ("珠海", 22.2707, 113.5767),
    ("佛山", 23.0215, 113.1214),
    ("东莞", 23.0207, 113.7518),
    ("西宁", 36.6171, 101.7782),
    ("银川", 38.4872, 106.2309),]

def seed_cities_if_empty(db: Session):
    if db.query(City).count() > 0:
        return
    for name, lat, lon in SEED_CITIES:
        db.add(City(name=name, lat=lat, lon=lon))
    db.commit()

@app.get("/cities", response_class=HTMLResponse)
def cities_page(request: Request, db: Session = Depends(get_db)):
    seed_cities_if_empty(db)
    cities = db.query(City).order_by(City.id.desc()).limit(300).all()
    return templates.TemplateResponse("cities.html", {"request": request, "cities": cities})

@app.post("/cities/upsert")
def cities_upsert(name: str = Form(...), lat: float = Form(...), lon: float = Form(...), db: Session = Depends(get_db)):
    nm = name.strip()
    rec = db.query(City).filter(City.name == nm).first()
    if rec:
        rec.lat = float(lat); rec.lon = float(lon)
    else:
        db.add(City(name=nm, lat=float(lat), lon=float(lon)))
    db.commit()
    return RedirectResponse(url="/cities", status_code=303)

@app.post("/cities/delete/{city_id}")
def cities_delete(city_id: int, db: Session = Depends(get_db)):
    rec = db.query(City).filter(City.id == city_id).first()
    if rec:
        db.delete(rec); db.commit()
    return RedirectResponse(url="/cities", status_code=303)

@app.post("/cities/import")
@app.post("/cities/import")
def cities_import(file: UploadFile, db: Session = Depends(get_db)):
    # CSV: name,lat,lon  (带表头也可以)
    import csv, io
    content = file.file.read()
    text = content.decode("utf-8-sig", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    for r in rows:
        if not r or len(r) < 3:
            continue
        if r[0].strip() in ("name","城市","city"):
            continue
        nm = r[0].strip()
        try:
            lat = float(r[1]); lon = float(r[2])
        except:
            continue
        rec = db.query(City).filter(City.name == nm).first()
        if rec:
            rec.lat = lat; rec.lon = lon
        else:
            db.add(City(name=nm, lat=lat, lon=lon))
    db.commit()
    return RedirectResponse(url="/cities", status_code=303)



@app.get("/imports", response_class=HTMLResponse)
def imports_page(request: Request):
    return templates.TemplateResponse("imports.html", {"request": request})

@app.get("/imports/template/auditors.xlsx")
def download_auditors_template():
    headers = ['姓名', '性别(男/女/不限)', '等级(A/B/C)', '可带队(是/否)', '常驻城市', '每周上限(院次)', '状态(在岗/请假/冻结)', '本月已排院次', '本月差旅天数', '连续工作天数', '上次结束城市', '上次结束日期(YYYY-MM-DD)']
    explain = ['必填', '可选', '必填', '可选', '必填', '默认1', '默认在岗', '默认0', '默认0', '默认0', '可空', '可空']
    example = [['张三', '男', 'A', '是', '北京', 1, '在岗', 0, 0, 0, '苏州', '2026-01-20'], ['李四', '女', 'B', '是', '上海', 2, '在岗', 0, 0, 0, '', '']]
    bio = make_xlsx_template(headers, [explain] + example, sheet_name="稽查员")
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": _content_disposition_utf8("稽查员模板.xlsx", "auditors.xlsx")})


@app.get("/imports/template/tasks.xlsx")
def download_tasks_template():
    headers = ['项目名称', '客户/申办方', '需要A带队(是/否)', '所需人数', '任务天数(用于推算结束)', '性别要求(男/女/不限)', '硬指定人员(可空)', '软指定专家/老师(可空)', '中心城市', '开始日期(YYYY-MM-DD)', '结束日期(可空)']
    explain = ['必填', '可空', '默认否', '默认1', '默认1', '默认不限', '可空', '可空(加分优先)', '必填', '必填', '可空(优先使用)']
    example = [['项目A', '申办方X', '否', 2, 2, '不限', '', '张三', '苏州', '2026-02-01', ''], ['项目B', '申办方Y', '是', 1, 3, '女', '', '', '北京', '2026-02-03', '2026-02-06']]
    bio = make_xlsx_template(headers, [explain] + example, sheet_name="任务")
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": _content_disposition_utf8("任务模板.xlsx", "tasks.xlsx")})


@app.post("/imports/upload/auditors")
def upload_auditors_xlsx(file: UploadFile = File(...), db: Session = Depends(get_db)):
    headers, rows = read_xlsx_rows(file)
    if not rows:
        raise HTTPException(status_code=400, detail="未读取到数据行")
    # map indices by header prefix
    ALIASES = {'name': ['姓名', 'name'], 'gender': ['性别(男/女/不限)', 'gender(男/女/不限)', 'gender'], 'group_level': ['等级(A/B/C)', 'group_level(A/B/C)', 'group_level'], 'can_lead_team': ['可带队(是/否)', 'can_lead_team(是/否)', 'can_lead_team'], 'base_city': ['常驻城市', 'base_city'], 'max_weekly_tasks': ['每周上限(院次)', 'max_weekly_tasks'], 'status': ['状态(在岗/请假/冻结)', 'status'], 'monthly_cases': ['本月已排院次', 'monthly_cases'], 'travel_days': ['本月差旅天数', 'travel_days'], 'continuous_days': ['连续工作天数', 'continuous_days'], 'last_task_end_city': ['上次结束城市', 'last_task_end_city'], 'last_task_end_date': ['上次结束日期(YYYY-MM-DD)', 'last_task_end_date(YYYY-MM-DD)', 'last_task_end_date']}
    def idx(key):
        for cand in ALIASES.get(key, [key]):
            for i,h in enumerate(headers):
                if str(h).strip() == str(cand).strip():
                    return i
        for i,h in enumerate(headers):
            hs=str(h).strip()
            for cand in ALIASES.get(key,[key]):
                if hs.startswith(str(cand).strip()):
                    return i
        return None
    import datetime as _dt
    from .models import Auditor
    n_ok = 0
    for r in rows:
        name = str(r[idx("name")] or "").strip()
        if not name:
            continue
        gender = str(r[idx("gender")] or "不限").strip()
        group_level = str(r[idx("group_level")] or "B").strip().upper()
        can_lead_team = str(r[idx("can_lead_team")] or "否").strip()
        base_city = str(r[idx("base_city")] or "").strip()
        max_weekly_tasks = int(r[idx("max_weekly_tasks")] or 1)
        status = str(r[idx("status")] or "在岗").strip()
        monthly_cases = int(r[idx("monthly_cases")] or 0)
        travel_days = int(r[idx("travel_days")] or 0)
        continuous_days = int(r[idx("continuous_days")] or 0)
        last_city = str(r[idx("last_task_end_city")] or "").strip() or None
        last_date_raw = r[idx("last_task_end_date")] if idx("last_task_end_date") is not None else None
        last_date = None
        if last_date_raw:
            if isinstance(last_date_raw, _dt.datetime):
                last_date = last_date_raw.date()
            elif isinstance(last_date_raw, _dt.date):
                last_date = last_date_raw
            else:
                try:
                    last_date = _dt.datetime.strptime(str(last_date_raw).strip(), "%Y-%m-%d").date()
                except:
                    last_date = None
        rec = db.query(Auditor).filter(Auditor.name == name).first()
        if rec:
            rec.gender = gender; rec.group_level = group_level
            rec.can_lead_team = (can_lead_team in ("是","Y","y","yes","YES","True","true","1"))
            rec.base_city = base_city
            rec.max_weekly_tasks = max_weekly_tasks
            rec.status = status
            rec.monthly_cases = monthly_cases
            rec.travel_days = travel_days
            rec.continuous_days = continuous_days
            rec.last_task_end_city = last_city
            rec.last_task_end_date = last_date
        else:
            db.add(Auditor(
                name=name, gender=gender, group_level=group_level,
                can_lead_team=(can_lead_team in ("是","Y","y","yes","YES","True","true","1")),
                base_city=base_city, max_weekly_tasks=max_weekly_tasks, status=status,
                monthly_cases=monthly_cases, travel_days=travel_days, continuous_days=continuous_days,
                last_task_end_city=last_city, last_task_end_date=last_date
            ))
        n_ok += 1
    db.commit()
    return RedirectResponse(url=f"/auditors?imported={n_ok}", status_code=303)

@app.post("/imports/upload/tasks")
def upload_tasks_xlsx(file: UploadFile = File(...), db: Session = Depends(get_db)):
    headers, rows = read_xlsx_rows(file)
    if not rows:
        raise HTTPException(status_code=400, detail="未读取到数据行")
    ALIASES = {'project_name': ['项目名称', 'project_name'], 'customer_name': ['客户/申办方', 'customer_name'], 'need_expert': ['需要A带队(是/否)', 'need_expert(是/否)', 'need_expert'], 'required_headcount': ['所需人数', 'required_headcount'], 'required_days': ['任务天数(用于推算结束)', 'required_days'], 'required_gender': ['性别要求(男/女/不限)', 'required_gender'], 'specified_auditors': ['硬指定人员(可空)', 'specified_auditors'], 'preferred_experts': ['软指定专家/老师(可空)', 'preferred_experts'], 'site_city': ['中心城市', 'site_city'], 'start_date': ['开始日期(YYYY-MM-DD)', 'start_date(YYYY-MM-DD)', 'start_date'], 'end_date': ['结束日期(可空)', 'end_date']}
    def idx(key):
        for cand in ALIASES.get(key, [key]):
            for i,h in enumerate(headers):
                if str(h).strip() == str(cand).strip():
                    return i
        for i,h in enumerate(headers):
            hs=str(h).strip()
            for cand in ALIASES.get(key,[key]):
                if hs.startswith(str(cand).strip()):
                    return i
        return None
    import datetime as _dt
    from .models import Task
    n_ok = 0
    for r in rows:
        project_name = str(r[idx("project_name")] or "").strip()
        if not project_name:
            continue
        customer_name = str(r[idx("customer_name")] or "").strip() or None
        need_expert = str(r[idx("need_expert")] or "否").strip() in ("是","Y","y","yes","YES","True","true","1")
        required_headcount = int(r[idx("required_headcount")] or 1)
        required_days = int(r[idx("required_days")] or 1)
        required_gender = str(r[idx("required_gender")] or "不限").strip()
        specified = str(r[idx("specified_auditors")] or "").strip() or None
        preferred = str(r[idx("preferred_experts")] or "").strip() or None
        site_city = str(r[idx("site_city")] or "").strip()
        sd_raw = r[idx("start_date")] if idx("start_date") is not None else None
        start_date = None
        end_date = None
        if sd_raw:
            if isinstance(sd_raw, _dt.datetime):
                start_date = sd_raw.date()
            elif isinstance(sd_raw, _dt.date):
                start_date = sd_raw
            else:
                start_date = _dt.datetime.strptime(str(sd_raw).strip(), "%Y-%m-%d").date()
        ed_raw = r[idx("end_date")] if idx("end_date") is not None else None
        if ed_raw:
            if isinstance(ed_raw, _dt.datetime):
                end_date = ed_raw.date()
            elif isinstance(ed_raw, _dt.date):
                end_date = ed_raw
            else:
                try:
                    end_date = _dt.datetime.strptime(str(ed_raw).strip(), "%Y-%m-%d").date()
                except:
                    end_date = None
        # upsert by project_name+start_date+site_city
        rec = db.query(Task).filter(Task.project_name==project_name, Task.start_date==start_date, Task.site_city==site_city).first()
        if rec:
            rec.customer_name = customer_name
            rec.need_expert = need_expert
            rec.required_headcount = required_headcount
            rec.required_days = required_days
            rec.required_gender = required_gender
            rec.specified_auditors = specified
            rec.preferred_experts = preferred
            rec.end_date = end_date
        else:
            db.add(Task(
                project_name=project_name, customer_name=customer_name, need_expert=need_expert,
                required_headcount=required_headcount, required_days=required_days, required_gender=required_gender,
                specified_auditors=specified, preferred_experts=preferred, site_city=site_city, start_date=start_date, end_date=end_date
            ))
        n_ok += 1
    db.commit()
    return RedirectResponse(url=f"/tasks?imported={n_ok}", status_code=303)


# ---- Schedule ----
@app.get("/schedule", response_class=HTMLResponse)
def schedule_page(request: Request, db: Session = Depends(get_db)):
    seed_city_distances_if_empty(db)
    tasks = db.query(Task).order_by(Task.id.desc()).all()
    schedules = db.query(Schedule).order_by(Schedule.id.desc()).limit(120).all()
    return templates.TemplateResponse("schedule.html", {"request": request, "tasks": tasks, "schedules": schedules,
                                                       "picked_task": None, "candidates": None, "team": None, "error": None})

@app.post("/schedule/recommend", response_class=HTMLResponse)
def schedule_recommend(request: Request, task_id: int = Form(...), db: Session = Depends(get_db)):
    task = db.query(Task).filter(Task.id == task_id).first()
    auditors = db.query(Auditor).all()
    schedules_all = db.query(Schedule).all()

    error = None; candidates = []; team = None
    if task:
        candidates = build_candidates(db, task, auditors, schedules_all)
        team = propose_team(task, candidates)
        if not team:
            error = "无可用团队方案：检查负责人/专家A组/人数不足/指定人冲突/每周上限/缓冲日冲突等。"

    tasks = db.query(Task).order_by(Task.id.desc()).all()
    schedules_recent = db.query(Schedule).order_by(Schedule.id.desc()).limit(120).all()
    return templates.TemplateResponse("schedule.html", {"request": request, "tasks": tasks, "schedules": schedules_recent,
                                                       "picked_task": task, "candidates": candidates[:25], "team": team, "error": error})

def assign_team(db: Session, task: Task, leader_id: int, member_ids):
    start_date = task.start_date
    end_date = task.start_date + timedelta(days=task.required_days - 1)

    def add_schedule(auditor_id: int, role: str):
        auditor = db.query(Auditor).filter(Auditor.id == auditor_id).first()
        if not auditor:
            return
        from_city = compute_from_city(auditor, task)
        km = get_distance_km(db, from_city, task.site_city)

        db.add(Schedule(
            task_id=task.id,
            auditor_id=auditor.id,
            role=role,
            start_date=start_date,
            end_date=end_date,
            travel_from_city=from_city,
            travel_to_city=task.site_city,
            distance_km=km,
            score=0.0,
            status="confirmed",
        ))

        auditor.monthly_cases += 1
        days = (end_date - start_date).days + 1
        auditor.travel_days += max(0, days)
        auditor.continuous_days = max(auditor.continuous_days, days)

        auditor.last_task_end_city = task.site_city
        auditor.last_task_end_date = end_date

    add_schedule(int(leader_id), "leader")
    for mid in member_ids:
        if int(mid) != int(leader_id):
            add_schedule(int(mid), "member")

@app.post("/schedule/assign_team")
def schedule_assign_team(task_id: int = Form(...), leader_id: int = Form(...), member_ids: str = Form(""), db: Session = Depends(get_db)):
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        return RedirectResponse(url="/schedule", status_code=303)
    ids = [i for i in re.split(r"[，,\s]+", member_ids.strip()) if i.strip()]
    mids = []
    for x in ids:
        try: mids.append(int(x))
        except: pass
    assign_team(db, task, int(leader_id), mids)
    db.commit()
    return RedirectResponse(url="/schedule", status_code=303)

@app.post("/schedule/delete/{schedule_id}")
def schedule_delete(schedule_id: int, db: Session = Depends(get_db)):
    s = db.query(Schedule).filter(Schedule.id == schedule_id).first()
    if s:
        db.delete(s); db.commit()
    return RedirectResponse(url="/schedule", status_code=303)

# ---- Batch scheduling ----
@app.get("/batch", response_class=HTMLResponse)
def batch_page(request: Request, db: Session = Depends(get_db)):
    seed_city_distances_if_empty(db)
    return templates.TemplateResponse("batch.html", {"request": request, "result": None})

@app.post("/batch/run", response_class=HTMLResponse)
def batch_run(request: Request, date_start: str = Form(...), date_end: str = Form(...), mode: str = Form("greedy"), db: Session = Depends(get_db)):
    seed_city_distances_if_empty(db)
    d1 = parse_date(date_start); d2 = parse_date(date_end)
    if d2 < d1:
        d1, d2 = d2, d1

    # 未排过的任务：task_id 不在 schedules 中
    scheduled_task_ids = {tid for (tid,) in db.query(Schedule.task_id).distinct().all()}
    tasks = db.query(Task).filter(Task.start_date >= d1, Task.start_date <= d2).all()
    tasks = [t for t in tasks if t.id not in scheduled_task_ids]

    # 排序：need_expert优先 > 人数多优先 > 开始日期早
    tasks.sort(key=lambda t: (0 if t.need_expert else 1, -t.required_headcount, t.start_date))

    auditors = db.query(Auditor).all()

    report = {"assigned": [], "skipped": []}

    for t in tasks:
        schedules_all = db.query(Schedule).all()  # 动态更新后重新取
        candidates = build_candidates(db, t, auditors, schedules_all)

        # 生成若干团队候选（负责人TOPK × 组员组合）
        team = propose_team(t, candidates)

        if mode == "optimized" and candidates:
            auditor_lookup = {a.id: a for a in auditors}
            # 计算当前平均月度院次（用于负荷均衡惩罚）
            avg_cases = float(sum(a.monthly_cases for a in auditors) / max(1, len(auditors)))
            # 批量内“本周使用次数”临时计数（同周更均衡）
            if "batch_week_counts" not in report:
                report["batch_week_counts"] = {}
            batch_week_counts = report["batch_week_counts"]

            # 通过枚举若干负责人/组员组合，挑选目标函数最小的团队
            # K1: 负责人候选取前5个可带队（need_expert=是则仅A）
            leader_pool = [c for c in candidates if c.can_lead_team]
            if t.need_expert:
                leader_pool = [c for c in leader_pool if c.group_level == "A"]
            leader_pool = leader_pool[:5]

            best_team = None
            best_obj = None

            # 组员候选池取前12
            member_pool_all = candidates[:12]

            for leader in leader_pool:
                member_pool = [c for c in member_pool_all if c.auditor_id != leader.auditor_id]
                need_n = max(0, int(t.required_headcount) - 1)
                if need_n == 0:
                    cand_team = type("Tmp", (), {})()
                    # 构造临时TeamProposal对象
                    from .scheduler import TeamProposal
                    cand_team = TeamProposal(leader=leader, members=[], team_score=leader.score, notes="optimized-single")
                    obj = team_objective(cand_team, auditor_lookup, avg_cases, batch_week_counts)
                    if best_obj is None or obj < best_obj:
                        best_obj, best_team = obj, cand_team
                    continue

                # 简化：直接取“member_pool按分数排序”中的前need_n作为一个团队方案
                # 再额外尝试几种“换人”组合（前need_n + 替换一个成员）
                base_members = member_pool[:need_n]
                from .scheduler import TeamProposal
                cand_team = TeamProposal(leader=leader, members=base_members, team_score=leader.score + sum(m.score for m in base_members)/max(1,len(base_members)), notes="optimized")
                obj = team_objective(cand_team, auditor_lookup, avg_cases, batch_week_counts)
                if best_obj is None or obj < best_obj:
                    best_obj, best_team = obj, cand_team

                # 尝试替换：用后面的候选替换base_members中的一个位置（最多尝试6次）
                extras = member_pool[need_n:need_n+6]
                for ex in extras:
                    if not base_members:
                        break
                    trial_members = base_members[:-1] + [ex]
                    cand_team2 = TeamProposal(leader=leader, members=trial_members, team_score=leader.score + sum(m.score for m in trial_members)/max(1,len(trial_members)), notes="optimized-swap")
                    obj2 = team_objective(cand_team2, auditor_lookup, avg_cases, batch_week_counts)
                    if best_obj is None or obj2 < best_obj:
                        best_obj, best_team = obj2, cand_team2

            if best_team:
                team = best_team

        if not team:
            report["skipped"].append({"task_id": t.id, "project": t.project_name, "reason": "无可用团队"})
            continue

        leader_id = team.leader.auditor_id
        member_ids = [m.auditor_id for m in team.members]
        assign_team(db, t, leader_id, member_ids)

        # 更新批量内本周使用次数（用于下一任务的均衡）
        if mode == "optimized":
            if "batch_week_counts" not in report:
                report["batch_week_counts"] = {}
            bwc = report["batch_week_counts"]
            for aid in [leader_id] + member_ids:
                bwc[aid] = int(bwc.get(aid, 0)) + 1

        db.commit()
        report["assigned"].append({"task_id": t.id, "project": t.project_name,
                                  "leader": team.leader.auditor_name,
                                  "members": [m.auditor_name for m in team.members]})

    return templates.TemplateResponse("batch.html", {"request": request, "result": report})